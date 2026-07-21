import sys, io
# Force UTF-8 output on Windows so Unicode characters print correctly.
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

"""
doc_extractor.py
================
Reads documents from ./input_docs, extracts invoice/receipt line-item data
using a local Ollama LLM, and writes structured JSON results to ./output_docs.

Supported file types
--------------------
  • PDF  — pdfplumber (text-based) → PyPDF2 fallback
  • PNG / JPG / JPEG / TIFF / BMP — pytesseract OCR
  • TXT  — plain read
"""

import os
import json
import re
import logging
from pathlib import Path
from typing import Optional

# ── Third-party ──────────────────────────────────────────────────────────────
try:
    import pdfplumber
except ImportError:
    pdfplumber = None  # type: ignore

try:
    import PyPDF2
except ImportError:
    PyPDF2 = None  # type: ignore

try:
    import pymupdf as fitz
except ImportError:
    try:
        import fitz  # type: ignore
    except ImportError:
        fitz = None  # type: ignore

try:
    import numpy as np
except ImportError:
    np = None

_paddle_ocr_engine = None
_paddle_ocr_initialized = False

def get_paddle_ocr():
    global _paddle_ocr_engine, _paddle_ocr_initialized
    if not _paddle_ocr_initialized:
        _paddle_ocr_initialized = True
        try:
            from paddleocr import PaddleOCR  # type: ignore # pylint: disable=import-error
            _paddle_ocr_engine = PaddleOCR(use_angle_cls=True, lang='en', show_log=False)
        except Exception:
            _paddle_ocr_engine = None
    return _paddle_ocr_engine

_easyocr_engine = None
_easyocr_initialized = False

def get_easy_ocr():
    global _easyocr_engine, _easyocr_initialized
    if not _easyocr_initialized:
        _easyocr_initialized = True
        try:
            import easyocr
            _easyocr_engine = easyocr.Reader(['en'])
        except Exception:
            _easyocr_engine = None
    return _easyocr_engine

try:
    import docx  # python-docx
except ImportError:
    docx = None  # type: ignore

try:
    import ollama
except ImportError:
    raise ImportError(
        "The 'ollama' package is required. Install it with:  pip install ollama"
    )

# ── Configuration ─────────────────────────────────────────────────────────────
OUTPUT_DIR  = Path("output_docs")

# Hybrid Cascade Configuration:
# Tier 1 (Fast Pass): Uses 3B model for 3x speed on initial extraction (fits 100% in 4GB VRAM)
OLLAMA_MODEL        = "qwen2.5:3b"
# Tier 2 (Heavy Escalation): Uses 7B model when math discrepancy is detected or during self-correction
OLLAMA_HEAVY_MODEL  = "qwen2.5:7b"
AI_ENGINE       = "ollama"
FALLBACK_ENGINE = "ollama"
MAX_WORKERS     = 4

SUPPORTED_EXTENSIONS = {".pdf", ".png", ".jpg", ".jpeg", ".tiff", ".tif", ".bmp", ".txt", ".docx"}

# ── Logging ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger(__name__)

# ─────────────────────────────────────────────────────────────────────────────
#  TEXT EXTRACTION
# ─────────────────────────────────────────────────────────────────────────────

def format_ocr_result(result) -> str:
    if not result:
        return ""
    
    items = []
    for bbox, text, prob in result:
        x_coords = [p[0] for p in bbox]
        y_coords = [p[1] for p in bbox]
        
        y_min, y_max = min(y_coords), max(y_coords)
        items.append({
            'text': text,
            'x_min': min(x_coords),
            'x_max': max(x_coords),
            'y_center': (y_min + y_max) / 2,
            'height': y_max - y_min
        })
        
    items.sort(key=lambda x: x['y_center'])
    
    lines = []
    current_line = []
    
    for item in items:
        if not current_line:
            current_line.append(item)
        else:
            # Check against the average y_center of the current line
            avg_y = sum(i['y_center'] for i in current_line) / len(current_line)
            # Use 0.5 height tolerance to group slightly misaligned items into the same row
            if abs(item['y_center'] - avg_y) <= item['height'] * 0.5:
                current_line.append(item)
            else:
                lines.append(current_line)
                current_line = [item]
                
    if current_line:
        lines.append(current_line)
        
    text_lines = []
    for line in lines:
        line.sort(key=lambda x: x['x_min'])
        line_str = ""
        last_x_max = None
        for i, it in enumerate(line):
            if i == 0:
                line_str += it['text']
            else:
                dist = it['x_min'] - last_x_max
                spaces = max(1, int(dist / 15)) if dist > 0 else 1
                line_str += (" " * spaces) + it['text']
            last_x_max = it['x_max']
        text_lines.append(line_str.strip())
        
    return "\n".join(text_lines)

def xycut_reading_order(words: list[dict], page_width: float = 0, page_height: float = 0) -> list[dict]:
    """
    Enhanced XY-Cut++ reading order algorithm.
    Recursively segments the page into layout blocks using X and Y projection gaps,
    and returns words ordered top-to-bottom, left-to-right within structural hierarchy.
    """
    if not words:
        return []
    if len(words) <= 1:
        return words

    def get_bbox(w):
        x0 = float(w.get("x0", 0))
        x1 = float(w.get("x1", x0))
        y0 = float(w.get("top", w.get("y0", 0)))
        y1 = float(w.get("bottom", w.get("y1", y0)))
        return x0, y0, x1, y1

    min_x = min(get_bbox(w)[0] for w in words)
    max_x = max(get_bbox(w)[2] for w in words)
    min_y = min(get_bbox(w)[1] for w in words)
    max_y = max(get_bbox(w)[3] for w in words)
    
    region_width = max_x - min_x
    region_height = max_y - min_y
    
    if region_height < 15 or region_width < 15 or len(words) <= 3:
        def sort_key(w):
            x0, y0, x1, y1 = get_bbox(w)
            return (round((y0 + y1) / 2.0 / 5.0) * 5.0, x0)
        return sorted(words, key=sort_key)

    def cut_recursive(sub_words: list[dict], x_min: float, y_min: float, x_max: float, y_max: float) -> list[dict]:
        if len(sub_words) <= 2:
            return sorted(sub_words, key=lambda w: (round((get_bbox(w)[1] + get_bbox(w)[3]) / 2.0 / 5.0) * 5.0, get_bbox(w)[0]))

        w_width = max(1.0, x_max - x_min)
        w_height = max(1.0, y_max - y_min)

        y_intervals = []
        for w in sub_words:
            bx0, by0, bx1, by1 = get_bbox(w)
            y_intervals.append((by0, by1, w))
        y_intervals.sort(key=lambda item: item[0])

        best_y_gap = 0.0
        best_y_cut = None
        current_y_max = y_intervals[0][1]
        for i in range(len(y_intervals) - 1):
            if y_intervals[i][1] > current_y_max:
                current_y_max = y_intervals[i][1]
            next_y_min = y_intervals[i + 1][0]
            if next_y_min > current_y_max:
                gap = next_y_min - current_y_max
                if gap > best_y_gap:
                    best_y_gap = gap
                    best_y_cut = (current_y_max + next_y_min) / 2.0

        x_intervals = []
        for w in sub_words:
            bx0, by0, bx1, by1 = get_bbox(w)
            x_intervals.append((bx0, bx1, w))
        x_intervals.sort(key=lambda item: item[0])

        best_x_gap = 0.0
        best_x_cut = None
        current_x_max = x_intervals[0][1]
        for i in range(len(x_intervals) - 1):
            if x_intervals[i][1] > current_x_max:
                current_x_max = x_intervals[i][1]
            next_x_min = x_intervals[i + 1][0]
            if next_x_min > current_x_max:
                gap = next_x_min - current_x_max
                if gap > best_x_gap:
                    best_x_gap = gap
                    best_x_cut = (current_x_max + next_x_min) / 2.0

        y_threshold = max(3.0, w_height * 0.008)
        x_threshold = max(12.0, w_width * 0.02)

        has_y_cut = best_y_cut is not None and best_y_gap >= y_threshold
        has_x_cut = best_x_cut is not None and best_x_gap >= x_threshold

        if not has_y_cut and not has_x_cut:
            def base_sort_key(w):
                bx0, by0, bx1, by1 = get_bbox(w)
                row_y = round((by0 + by1) / 2.0 / 6.0) * 6.0
                return (row_y, bx0)
            return sorted(sub_words, key=base_sort_key)

        if has_y_cut:
            top_words = [w for w in sub_words if (get_bbox(w)[1] + get_bbox(w)[3])/2.0 < best_y_cut]
            bot_words = [w for w in sub_words if (get_bbox(w)[1] + get_bbox(w)[3])/2.0 >= best_y_cut]
            if top_words and bot_words:
                return cut_recursive(top_words, x_min, y_min, x_max, best_y_cut) + cut_recursive(bot_words, x_min, best_y_cut, x_max, y_max)
        if has_x_cut:
            left_words = [w for w in sub_words if (get_bbox(w)[0] + get_bbox(w)[2])/2.0 < best_x_cut]
            right_words = [w for w in sub_words if (get_bbox(w)[0] + get_bbox(w)[2])/2.0 >= best_x_cut]
            if left_words and right_words:
                return cut_recursive(left_words, x_min, y_min, best_x_cut, y_max) + cut_recursive(right_words, best_x_cut, y_min, x_max, y_max)

        return sorted(sub_words, key=lambda w: (round((get_bbox(w)[1] + get_bbox(w)[3]) / 2.0 / 6.0) * 6.0, get_bbox(w)[0]))

    return cut_recursive(words, min_x, min_y, max_x, max_y)

def extract_text_with_xycut(page) -> str:
    """
    Extracts text from a pdfplumber page using word bounding boxes and the XY-Cut++ algorithm.
    Groups ordered words into lines and preserves spatial spacing.
    """
    try:
        words = page.extract_words()
        if not words:
            return ""
        
        ordered_words = xycut_reading_order(words, getattr(page, "width", 0), getattr(page, "height", 0))
        
        items = []
        for w in ordered_words:
            x0 = float(w.get("x0", 0))
            x1 = float(w.get("x1", x0))
            y0 = float(w.get("top", w.get("y0", 0)))
            y1 = float(w.get("bottom", w.get("y1", y0)))
            items.append({
                'text': w.get("text", ""),
                'x_min': x0,
                'x_max': x1,
                'y_center': (y0 + y1) / 2.0,
                'height': max(1.0, y1 - y0)
            })
            
        lines = []
        current_line = []
        for item in items:
            if not current_line:
                current_line.append(item)
            else:
                avg_y = sum(i['y_center'] for i in current_line) / len(current_line)
                if abs(item['y_center'] - avg_y) <= item['height'] * 0.55:
                    current_line.append(item)
                else:
                    lines.append(current_line)
                    current_line = [item]
        if current_line:
            lines.append(current_line)
            
        text_lines = []
        for line in lines:
            line.sort(key=lambda x: x['x_min'])
            line_str = ""
            last_x_max = None
            for i, it in enumerate(line):
                if i == 0:
                    line_str += it['text']
                else:
                    dist = it['x_min'] - last_x_max
                    spaces = max(1, int(dist / 6)) if dist > 0 else 1
                    line_str += (" " * spaces) + it['text']
                last_x_max = it['x_max']
            text_lines.append(line_str.strip())
            
        return "\n".join(text_lines)
    except Exception as exc:
        log.warning("  [XY-Cut] Failed to extract words with XY-Cut++: %s", exc)
        return ""

def extract_text_from_pdf(file_path: Path) -> list[str]:
    """
    Extract text from a PDF file page by page.
    """
    if pdfplumber is not None:
        try:
            with pdfplumber.open(file_path) as pdf:
                pages_text = []
                for p in pdf.pages:
                    text = p.extract_text(layout=True) if hasattr(p, 'extract_text') else ""
                    if not text or len(text.strip()) < 10:
                        text = extract_text_with_xycut(p)
                    if not text or len(text.strip()) < 10:
                        text = p.extract_text(layout=False) or ""
                    pages_text.append(text)
                if any(pt.strip() for pt in pages_text):
                    return pages_text
                else:
                    log.info("  [PDF] pdfplumber found no text (likely scanned images) — trying PyPDF2 or OCR")
        except Exception as exc:
            log.warning("  [PDF] pdfplumber failed: %s — trying PyPDF2", exc)

    if PyPDF2 is not None:
        try:
            with open(file_path, "rb") as fh:
                reader = PyPDF2.PdfReader(fh)
                pages_text = []
                for page in reader.pages:
                    page_text = page.extract_text()
                    pages_text.append(page_text or "")
                if any(pages_text):
                    return pages_text
        except Exception as exc:
            log.warning("  [PDF] PyPDF2 also failed: %s", exc)

    if fitz is not None and np is not None:
        p_engine = get_paddle_ocr()
        if p_engine is not None:
            log.info("  [PDF] Text layer missing or very small, falling back to PaddleOCR...")
            try:
                doc = fitz.open(file_path)
                pages_text = []
                for page in doc:
                    pix = page.get_pixmap(dpi=250)
                    # Convert PyMuPDF pixmap to numpy array
                    img_array = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.height, pix.width, pix.n)
                    if pix.n == 4:
                        img_array = img_array[:, :, :3] # Convert RGBA to RGB
                    
                    result = p_engine.ocr(img_array, cls=True)
                    
                    if result and result[0]:
                        ocr_data = []
                        for line in result[0]:
                            bbox = line[0]
                            text = line[1][0]
                            prob = line[1][1]
                            ocr_data.append((bbox, text, prob))
                        page_text = format_ocr_result(ocr_data)
                        pages_text.append(page_text)
                    else:
                        pages_text.append("")
                        
                if any(pages_text):
                    return pages_text
            except Exception as exc:
                log.warning("  [PDF] PyMuPDF PaddleOCR fallback failed: %s — trying EasyOCR", exc)

        e_engine = get_easy_ocr()
        if e_engine is not None:
            log.info("  [PDF] Text layer missing or very small, falling back to EasyOCR...")
            try:
                doc = fitz.open(file_path)
                pages_text = []
                for page in doc:
                    pix = page.get_pixmap(dpi=250)
                    # Convert PyMuPDF pixmap to numpy array
                    img_array = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.height, pix.width, pix.n)
                    if pix.n == 4:
                        img_array = img_array[:, :, :3] # Convert RGBA to RGB
                    
                    result = e_engine.readtext(img_array)
                    
                    if result:
                        page_text = format_ocr_result(result)
                        pages_text.append(page_text)
                    else:
                        pages_text.append("")
                        
                if any(pages_text):
                    return pages_text
            except Exception as exc:
                log.warning("  [PDF] PyMuPDF EasyOCR fallback failed: %s", exc)

    log.error("  [PDF] Could not extract text from %s (or no OCR engine is available)", file_path.name)
    return []

def extract_text_from_image(file_path: Path) -> list[str]:
    """Extract text from an image using PaddleOCR (preferred) or EasyOCR (fallback)."""
    p_engine = get_paddle_ocr()
    if p_engine is not None:
        try:
            result = p_engine.ocr(str(file_path), cls=True)
            if result and result[0]:
                ocr_data = []
                for line in result[0]:
                    bbox = line[0]
                    text = line[1][0]
                    prob = line[1][1]
                    ocr_data.append((bbox, text, prob))
                page_text = format_ocr_result(ocr_data)
                return [page_text]
            return [""]
        except Exception as exc:
            log.warning("  [IMG] PaddleOCR failed for %s: %s — trying EasyOCR", file_path.name, exc)

    e_engine = get_easy_ocr()
    if e_engine is not None:
        try:
            result = e_engine.readtext(str(file_path))
            if result:
                text = format_ocr_result(result)
                return [text]
            return [""]
        except Exception as exc:
            log.error("  [IMG] EasyOCR failed for %s: %s", file_path.name, exc)
            return []
            
    log.error("  [IMG] No OCR engine is available (PaddleOCR and EasyOCR are both missing).")
    return []

def extract_text_from_txt(file_path: Path) -> list[str]:
    try:
        text = file_path.read_text(encoding="utf-8", errors="replace")
        return [text.strip()]
    except Exception as exc:
        log.error("  [TXT] Could not read %s: %s", file_path.name, exc)
        return []

def extract_text_from_docx(file_path: Path) -> list[str]:
    if docx is None:
        log.error("  [DOCX] python-docx not installed.")
        return []
    try:
        document = docx.Document(str(file_path))
        paragraphs = [p.text for p in document.paragraphs if p.text.strip()]
        text = "\n".join(paragraphs)
        return [text.strip()]
    except Exception as exc:
        log.error("  [DOCX] Failed to read %s: %s", file_path.name, exc)
        return []

def extract_text(file_path: Path) -> list[str]:
    """Dispatch to the correct extractor based on file extension."""
    ext = file_path.suffix.lower()
    pages = []
    if ext == ".pdf":
        pages = extract_text_from_pdf(file_path)
    elif ext in {".png", ".jpg", ".jpeg", ".tiff", ".tif", ".bmp"}:
        pages = extract_text_from_image(file_path)
    elif ext == ".txt":
        pages = extract_text_from_txt(file_path)
    elif ext == ".docx":
        pages = extract_text_from_docx(file_path)
    else:
        log.warning("  Unsupported extension '%s' for %s", ext, file_path.name)
        
    # Clean text
    cleaned_pages = []
    for p in pages:
        if p.strip():
            # Split merged category headers from serial numbers and items
            cat_headers = [
                r"ADMINISTRATIVE CHARGES", r"ADMISSION", r"BED CHARGES", r"DIETETICS DEPARTMENT",
                r"DOCTORS VISIT CHARGE", r"LABORATORY", r"NURSING CHARGES", r"PROCEDURE/SERVICE CHARGE",
                r"RADIOLOGY", r"CONSUMABLES-IP PHARMACY", r"MEDICINE-IP PHARMACY", r"MATERIALS",
                r"IMPLANT CHARGES", r"DIET", r"GST ON BED CHARGES \d+%", r"THEATRE CHARGES",
                r"SURGEON FEE", r"ANASTHETISTS FEES", r"SURGICAL SUPPORT FEES", r"CSSD CHARGES",
                r"MISCELLANEOUS", r"PHARMACY DRUGS", r"PARTICULARS CHARGES"
            ]
            cat_pattern = r"^(" + "|".join(cat_headers) + r")\s+(\d+)\b"
            p = re.sub(cat_pattern, r"\1\n\2", p, flags=re.IGNORECASE | re.MULTILINE)

            # Remove "Location:" lines which the LLM often misinterprets as items
            p = re.sub(r"Location:.*", "", p)
            # Fix stray quotes inside numbers (e.g. 2'13.00 -> 213.00)
            p = re.sub(r'\b(\d+)[\'’](\d+\.\d{2})\b', r'\1\2', p)
            # Fix OCR misreads inside prices: [l I] -> 1 and [S s] -> 5 surrounded by digits
            p = re.sub(r'\b([0-9]+)[lI]([0-9]*\.[0-9]{2})\b', r'\g<1>1\2', p)
            p = re.sub(r'(?<=\d)[Ss](?=\d|\.\d{2})', '5', p)
            # Repair broken decimals separated by whitespace (e.g. 1234 . 00 -> 1234.00)
            p = re.sub(r'\b(\d+)\s*\.\s*(\d{2})\b', r'\1.\2', p)
            # Fix OCR misread where zeroes in prices are misread as o or O (like 6oo.oo -> 600.00)
            p = re.sub(r'(?<=\d)[oO]+(?:\.[oO0]{2})\b', lambda m: '0'*len(m.group(0).split('.')[0]) + '.00', p)
            # Fix OCR misread where .oo or .OO is misread instead of .00
            p = re.sub(r'\b(\d+)\.[oO]{2}\b', r'\1.00', p)
            # Fix OCR misread L000 or l000 for 1.000
            p = re.sub(r'\b[Ll]000\b', '1.000', p)
            # Fix missing decimal point for perfectly whole numbers ending in space+00 (like 120 00)
            p = re.sub(r'\b(\d+)\s+00\b', r'\1.00', p)
            # Separate table headers merged with first row on same line
            p = re.sub(r'\b(PARTICULARS\s+(?:CHARGES|AMOUNT|RATE|PRICE|QTY|QUANTITY))\s+(?=[A-Z0-9\[{])', r'\1\n', p, flags=re.IGNORECASE)
            # Strip purely patient metadata header lines
            p = re.sub(r'^.*?(?:\bCompany\b|\bMRD\s*No\b|\bPatient\s*Name\b|\bVisit\s*Code\b|\bVlslt\b|\bPatlent\b|\bAge,\s*Sex\b)[^\n]*$', '', p, flags=re.IGNORECASE | re.MULTILINE)
            # Strip patient metadata prefixes up to item code
            p = re.sub(r'^(?:.*?(?:\bCompany\b|\bMRD\s*No\b|\bPatient\s*Name\b|\bVisit\s*Code\b|\bVlslt\b|\bPatlent\b)[^\n]*?)(?=\b[A-Z]{2}-\d{2}-\d{4}\b)', '', p, flags=re.IGNORECASE | re.MULTILINE)
            # Clean blue ink handwriting OCR noise on summary lines
            p = re.sub(r'[\{\}\(\)\!]+|\b(?:CCH|sl|o)\b', '', p)
            p = re.sub(r'\b[T]\b\s+(?=OTHER CHARGES)', '', p)
            # Remove stray bullet points
            p = re.sub(r'^[o•]\s+', '', p, flags=re.MULTILINE)
            # Split horizontally merged line items (price followed by next item description)
            p = re.sub(r'(\d+\.\d{2})\s+(?=[A-Z][A-Z\s]+(?:\b\d|\b[A-Z]{2,}))', r'\1\n', p)
            # Clean up stray trailing page numbers/digits after prices at line ends
            p = re.sub(r'(\.\d{2})\s+\d+\s*$', r'\1', p, flags=re.MULTILINE)
            
            # Smartly merge split line items (e.g. medical devices split across two lines)
            # If the current line has a price at the end, but the PREVIOUS line has NO price, merge them.
            lines = p.split('\n')
            merged_lines = []
            for line in lines:
                if not line.strip(): continue
                line_has_price = bool(re.search(r'\d+(?:\.\d{2})?\s*$', line.strip()))
                
                if merged_lines:
                    prev_has_price = bool(re.search(r'\d+(?:\.\d{2})?\s*$', merged_lines[-1].strip()))
                    if not prev_has_price and line_has_price:
                        merged_lines[-1] = merged_lines[-1].strip() + " " + line.strip()
                        continue
                merged_lines.append(line)
            
            p = "\n".join(merged_lines)
            cleaned_pages.append(p.strip())
        else:
            cleaned_pages.append("")
    return cleaned_pages

# ─────────────────────────────────────────────────────────────────────────────
#  OLLAMA LLM EXTRACTION
# ─────────────────────────────────────────────────────────────────────────────

SYSTEM_PROMPT = """\
You are a JSON-only data extraction API.
You ONLY output raw JSON. You NEVER write explanations, descriptions, or prose.
If you cannot find line items, output: {{"page_number": %d, "items": []}}
Do NOT describe the document. Do NOT use markdown. Output ONLY the JSON object.
"""

USER_PROMPT_TEMPLATE = """\
Extract every single financial/billing line item from the invoice/receipt text below.

For each line item, extract the data and map it to the following standardized JSON keys:
- "Particulars": The primary item description, name, or service details.
- "Quantity": The number of units, quantity, count, or visits. If not specified, use 1.
- "Price": The unit price, rate, or charge per item.
- "NetAmt": The net amount, total charges, or row total for this item.

CRITICAL COLUMN ALIGNMENT RULE:
Identify the headers of the table columns in the text:
1. If the table columns are: [Sl# Description Date Qty Rate Gross Amount Discount]
   - Map "Quantity" to Qty (e.g. 5, 4, 1, etc.).
   - Map "Price" to Rate (e.g. 12.00, 201.00, etc.).
   - Map "NetAmt" to Gross Amount (e.g. 60.00, 201.00, etc.). Do NOT map NetAmt to the Discount column!
2. If the table columns are: [Sl# Particulars Qty Price Amount]
   - Map "Quantity" to Qty.
   - Map "Price" to Price.
   - Map "NetAmt" to Amount.
3. If the table columns only have [Particulars Quantity Amount] or similar (meaning there is NO unit price/rate column printed):
   - Map "Quantity" to Quantity/Qty.
   - Map "NetAmt" to Amount (row total).
   - Map "Price" to NetAmt / Quantity (calculate this unit price mathematically!).
4. Verify that for every extracted item: Price * Quantity is mathematically close to NetAmt. If not, check if you misaligned the columns!

OUTPUT FORMAT — respond with ONLY this JSON structure, nothing else:
{{
  "page_number": {page_num},
  "items": [
    {{"Particulars": "Item Description", "Quantity": 1.0, "Price": 150.0, "NetAmt": 150.0}},
    {{"Particulars": "Another Item", "Quantity": 2.0, "Price": 100.0, "NetAmt": 200.0}}
  ]
}}

RULES:
- Only extract FINANCIAL billing line items. If a page or a list contains NO financial prices (e.g. a medical report, clinical notes, patient details), IGNORE IT completely and return an empty list for "items".
- Do NOT extract patient metadata (Name, Age, Sex, Address, Dates, etc.) as line items.
- Extract the complete item description exactly as it appears. Do not truncate.
- All numeric fields must be plain numbers (no $ signs, no commas).
- Do NOT collapse or deduplicate repeating items. If the document has multiple identical lines (for example, multiple identical 'Injection' or 'Ward Charge' rows), you must output each one as a separate distinct item in the JSON list. Never skip or aggregate repeating charges.
- Do NOT extract category subtotal lines or summary rows (such as lines ending with 'Sub Total: ...' or generic rows named 'Order Item' that just repeat a section sum). Extract ONLY individual itemized charges.
- Do NOT extract Services Accounting Codes (SAC) or HSN codes (typically formatted as "SAC:XXXXXX" or "HSN:XXXXXX" or as 6-digit integers starting with 99) as prices or line items. They are tax classifications, NOT financial amounts.
- You MUST use the EXACT numeric values printed in the document. Never change, guess, or synthesize numbers. If the text says 4,620.00, you must output 4620.00. Outputting a slightly different number (like 4626.00 or 514.00) is a critical error and will cause the item to be discarded.
{ocr_instructions}

DO NOT write any sentence or paragraph. START your response with the {{ character.

--- INVOICE TEXT ---
{raw_text}
--- END ---
"""

RETRY_PROMPT_TEMPLATE = """\
Your previous response was not valid JSON. Try again.

Respond with ONLY a JSON object. Start immediately with {{ and end with }}.
No explanation. No markdown. No prose. Only JSON.

Schema:
{{"page_number": {page_num}, "items": [{{"Particulars": "...", "Quantity": 1.0, "Price": 0.0, "NetAmt": 0.0}}]}}

Invoice text:
{raw_text}
"""


def clean_json_response(raw: str) -> str:
    raw = raw.strip()
    fence_re = re.compile(r"```(?:json)?\s*([\s\S]*?)\s*```", re.MULTILINE)
    m = fence_re.search(raw)
    if m:
        return m.group(1).strip()

    brace_re = re.compile(r"(\{[\s\S]*\})", re.MULTILINE)
    m2 = brace_re.search(raw)
    if m2:
        raw = m2.group(1).strip()

    raw = raw.replace(r"\_", "_")
    return raw


def _call_llm(messages: list, label: str, model_override: Optional[str] = None) -> Optional[str]:
    # Schema definition for strict JSON output enforcement
    schema_def = {
        "type": "OBJECT",
        "properties": {
            "page_number": {"type": "INTEGER"},
            "items": {
                "type": "ARRAY",
                "items": {
                    "type": "OBJECT",
                    "properties": {
                        "Particulars": {"type": "STRING"},
                        "Quantity": {"type": "NUMBER"},
                        "Price": {"type": "NUMBER"},
                        "NetAmt": {"type": "NUMBER"}
                    },
                    "required": ["Particulars", "Quantity", "Price", "NetAmt"],
                    "additionalProperties": False
                }
            }
        },
        "required": ["page_number", "items"],
        "additionalProperties": False
    }

    try:
        if AI_ENGINE == "gemini":
            log.info("  [LLM] Calling Gemini Cloud API (%s)…", label)
            import urllib.request, json as _json, os
            api_key = os.environ.get("GEMINI_API_KEY", "")
            if not api_key:
                log.warning("  [GEMINI] GEMINI_API_KEY environment variable not set! Falling back to %s...", FALLBACK_ENGINE)
                return _call_fallback(messages, label, model_override=model_override)
            url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={api_key}"
            prompt_text = "\n\n".join(m["content"] for m in messages)
            payload = _json.dumps({
                "contents": [{"parts": [{"text": prompt_text}]}],
                "generationConfig": {
                    "responseMimeType": "application/json",
                    "responseSchema": schema_def
                }
            }).encode("utf-8")
            req = urllib.request.Request(url, data=payload, headers={"Content-Type": "application/json"})
            import time as _time
            for _attempt in range(5):
                try:
                    with urllib.request.urlopen(req) as resp:
                        res_data = _json.loads(resp.read().decode("utf-8"))
                        reply_text = res_data["candidates"][0]["content"]["parts"][0]["text"]
                        _time.sleep(4.2)
                        return reply_text
                except urllib.error.HTTPError as he:
                    if he.code == 429:
                        log.warning("  [GEMINI] Rate limit (429) hit. Pausing 15s to clear window...")
                        _time.sleep(15)
                    else:
                        log.warning("  [GEMINI] HTTP error %s. Falling back to %s...", he, FALLBACK_ENGINE)
                        return _call_fallback(messages, label, model_override=model_override)
            log.warning("  [GEMINI] Retries exhausted. Falling back to %s...", FALLBACK_ENGINE)
            return _call_fallback(messages, label, model_override=model_override)
        elif AI_ENGINE == "openai":
            log.info("  [LLM] Calling OpenAI Cloud API (%s)…", label)
            import urllib.request, json as _json, os
            api_key = os.environ.get("OPENAI_API_KEY", "")
            if not api_key:
                log.warning("  [OPENAI] OPENAI_API_KEY environment variable not set! Falling back to %s...", FALLBACK_ENGINE)
                return _call_fallback(messages, label, model_override=model_override)
            url = "https://api.openai.com/v1/chat/completions"
            # Lowercase types for OpenAI JSON Schema
            openai_schema = {
                "name": "extraction_schema",
                "strict": True,
                "schema": {
                    "type": "object",
                    "properties": {
                        "page_number": {"type": "integer"},
                        "items": {
                            "type": "array",
                            "items": {
                                "type": "object",
                                "properties": {
                                    "Particulars": {"type": "string"},
                                    "Quantity": {"type": "number"},
                                    "Price": {"type": "number"},
                                    "NetAmt": {"type": "number"}
                                },
                                "required": ["Particulars", "Quantity", "Price", "NetAmt"],
                                "additionalProperties": False
                            }
                        }
                    },
                    "required": ["page_number", "items"],
                    "additionalProperties": False
                }
            }
            payload = _json.dumps({
                "model": "gpt-4o-mini",
                "messages": messages,
                "response_format": {"type": "json_schema", "json_schema": openai_schema},
                "temperature": 0
            }).encode("utf-8")
            req = urllib.request.Request(url, data=payload, headers={"Content-Type": "application/json", "Authorization": f"Bearer {api_key}"})
            try:
                with urllib.request.urlopen(req) as resp:
                    res_data = _json.loads(resp.read().decode("utf-8"))
                    return res_data["choices"][0]["message"]["content"]
            except Exception as exc:
                log.warning("  [OPENAI] Call failed (%s). Falling back to %s...", exc, FALLBACK_ENGINE)
                return _call_fallback(messages, label, model_override=model_override)

        return _call_ollama_engine(messages, label, model_override=model_override)
    except Exception as exc:
        log.error("  [LLM] %s call failed: %s", AI_ENGINE, exc)
        if AI_ENGINE != FALLBACK_ENGINE:
            log.info("  [LLM] Attempting fallback engine '%s'...", FALLBACK_ENGINE)
            return _call_fallback(messages, label, model_override=model_override)
        return None


def _call_fallback(messages: list, label: str, model_override: Optional[str] = None) -> Optional[str]:
    if FALLBACK_ENGINE == "ollama":
        return _call_ollama_engine(messages, label + " [fallback]", model_override=model_override)
    return None


def _call_ollama_engine(messages: list, label: str, model_override: Optional[str] = None) -> Optional[str]:
    try:
        model_to_use = model_override if model_override else OLLAMA_MODEL
        log.info("  [LLM] Calling Ollama model '%s' (%s)…", model_to_use, label)
        response = ollama.chat(
            model=model_to_use,
            messages=messages,
            format="json",
            options={"temperature": 0, "seed": 42, "num_ctx": 4096, "num_predict": 2048},
        )
        reply: str = response["message"]["content"]
        log.info("  [LLM] Received %d chars from model", len(reply))
        return reply
    except Exception as exc:
        log.error("  [OLLAMA] Call failed: %s", exc)
        return None


def _parse_reply(raw_reply: str, page_num: int) -> Optional[dict]:
    cleaned = clean_json_response(raw_reply)
    try:
        data = json.loads(cleaned)
        if isinstance(data, dict):
            # Fix hallucinated schema keys if LLM returned "results" instead of "items"
            if "results" in data and "items" not in data:
                data["items"] = data.pop("results")
                
            # Fallback if there's any list in the dict
            if "items" not in data:
                for k, v in data.items():
                    if isinstance(v, list):
                        data["items"] = v
                        break
            
            if "items" not in data:
                data["items"] = []
                
            data["page_number"] = page_num
            return data
    except json.JSONDecodeError:
        pass
    return None


def get_case_insensitive(d: dict, keys: list[str], default=None):
    if not isinstance(d, dict):
        return default
    for k in keys:
        for dk in d:
            if dk.lower() == k.lower():
                return d[dk]
    return default


def detect_page_headers(page_text: str) -> list[str]:
    text_lower = page_text.lower()
    
    # Check for the 8-column header (Description + Cpt + Code + Discount)
    if "description" in text_lower and "cpt" in text_lower and "code" in text_lower and "discount" in text_lower:
        sl_key = "SI#" if "si#" in text_lower else "Sl#"
        return [sl_key, "Description", "Cpt Code", "Date", "Qty", "Rate", "Gross Amount", "Discount"]
        
    # Default 4-column header
    return ["Particulars", "Quantity", "Price", "NetAmt"]


def denormalize_item(item: dict, headers: list[str]) -> dict:
    denorm = {}
    for h in headers:
        h_lower = h.lower()
        if h_lower in ["sl#", "si#"]:
            val = get_case_insensitive(item, ["sl#", "si#"], "")
            denorm[h] = val
        elif h_lower == "description":
            denorm[h] = item.get("Particulars", get_case_insensitive(item, ["description"], ""))
        elif h_lower == "cpt code":
            denorm[h] = get_case_insensitive(item, ["cpt code", "cptcode"], "")
        elif h_lower == "date":
            denorm[h] = get_case_insensitive(item, ["date"], "")
        elif h_lower == "qty":
            denorm[h] = item.get("Quantity", item.get("Qty", 1.0))
        elif h_lower == "rate":
            denorm[h] = item.get("Price", item.get("Rate", 0.0))
        elif h_lower == "gross amount":
            denorm[h] = item.get("NetAmt", item.get("Gross Amount", 0.0))
        elif h_lower == "discount":
            denorm[h] = get_case_insensitive(item, ["discount", "disc"], 0.0)
        elif h_lower == "particulars":
            denorm[h] = item.get("Particulars", get_case_insensitive(item, ["description"], ""))
        elif h_lower == "quantity":
            denorm[h] = item.get("Quantity", item.get("Qty", 1.0))
        elif h_lower == "price":
            denorm[h] = item.get("Price", item.get("Rate", 0.0))
        elif h_lower == "netamt":
            denorm[h] = item.get("NetAmt", item.get("Gross Amount", 0.0))
        else:
            denorm[h] = item.get(h, "")
            
    # Clean numeric fields
    for h in headers:
        if h in ["Qty", "Quantity"]:
            try:
                denorm[h] = float(str(denorm[h]).replace(",", "").strip())
            except ValueError:
                denorm[h] = 1.0
        elif h in ["Rate", "Price", "Gross Amount", "NetAmt", "Discount"]:
            try:
                denorm[h] = float(str(denorm[h]).replace(",", "").strip())
            except ValueError:
                denorm[h] = 0.0
                
    return denorm


def normalize_item(item: dict) -> dict:
    if not isinstance(item, dict):
        return {}
    
    # 1. Particulars
    particulars = get_case_insensitive(item, ["Particulars", "Description", "Name", "Item", "Particular", "Text"], "")
    if particulars is None:
        particulars = ""
    else:
        particulars = str(particulars).strip()
        
    # 2. Quantity
    qty_val = get_case_insensitive(item, ["Quantity", "Qty", "Count", "No"], None)
    if qty_val in [None, "null", "", "None"]:
        qty = 1.0
    else:
        try:
            qty = float(str(qty_val).replace(",", "").strip())
        except ValueError:
            qty = 1.0
            
    # 3. Price
    price_val = get_case_insensitive(item, ["Price", "Rate", "Charges", "Unit Price", "UnitPrice"], None)
    price = None
    if price_val not in [None, "null", "", "None"]:
        try:
            price = float(str(price_val).replace(",", "").strip())
        except ValueError:
            pass
            
    # 4. NetAmt
    net_val = get_case_insensitive(item, ["NetAmt", "Amount", "Gross Amount", "NetAmount", "Total", "Charges Total"], None)
    net = None
    if net_val not in [None, "null", "", "None"]:
        try:
            net = float(str(net_val).replace(",", "").strip())
        except ValueError:
            pass
            
    # Fallbacks to align them mathematically
    if (price is None or price == 0.0) and (net is None or net == 0.0) and qty > 20:
        net = float(qty)
        price = float(qty)
        qty = 1.0
    elif price is None and net is not None:
        price = round(net / qty, 2) if qty > 0 else net
    elif net is None and price is not None:
        net = round(price * qty, 2)
    elif price is None and net is None:
        price = 0.0
        net = 0.0
        
    # Keep ALL other keys (e.g. Sl#, Date, Discount, Cpt Code) in the normalized dict!
    normalized = {}
    for k, v in item.items():
        k_lower = k.lower()
        if k_lower not in ["particulars", "description", "name", "item", "particular", "text",
                           "quantity", "qty", "count", "no",
                           "price", "rate", "charges", "unit price", "unitprice",
                           "netamt", "amount", "gross amount", "netamount", "total", "charges total", "item_type"]:
            normalized[k] = v
            
    normalized["Particulars"] = particulars
    normalized["Quantity"] = qty
    normalized["Qty"] = qty
    normalized["Price"] = price
    normalized["Rate"] = price
    normalized["NetAmt"] = net
    normalized["Gross Amount"] = net
    
    if bool(re.search(r'\b(CGST|SGST|IGST|GST|VAT)\b', particulars.upper())):
        normalized["item_type"] = "tax"
    else:
        normalized["item_type"] = "charge"
    
    return normalized


def find_matching_line(particulars, page_text):
    if not particulars or not page_text:
        return ""
    part_lower = particulars.lower().strip()
    
    # 1. Exact string match after removing trailing numbers/prices
    for line in page_text.splitlines():
        line_lower = line.lower().strip()
        if not line_lower:
            continue
        clean_text = re.sub(r'[\d\.\s,-]+$', '', line_lower).strip()
        if clean_text == part_lower:
            return line

    # 2. Substring match, picking the closest in text length
    substring_matches = []
    for line in page_text.splitlines():
        line_lower = line.lower().strip()
        if part_lower in line_lower:
            substring_matches.append(line)
    if substring_matches:
        return min(substring_matches, key=lambda l: abs(len(re.sub(r'[\d\.\s,-]+$', '', l.lower().strip()).strip()) - len(part_lower)))

    best_line = ""
    best_score = 0.0
    import difflib
    for line in page_text.splitlines():
        line_lower = line.lower().strip()
        if not line_lower:
            continue
        ratio = difflib.SequenceMatcher(None, part_lower, line_lower).ratio()
        if ratio > best_score:
            best_score = ratio
            best_line = line
            
    if best_score >= 0.5:
        return best_line
    return ""


def fix_ocr_json_math(items, page_text=""):
    """
    Given a list of item dictionaries, ensure that Qty * Price == NetAmt.
    If they don't match, recalculate NetAmt based on Qty and Price.
    """
    normalized = []
    for it in items:
        if isinstance(it, dict):
            normalized.append(normalize_item(it))
            
    merged_items = []
    i = 0
    while i < len(normalized):
        curr = normalized[i]
        
        if i + 1 < len(normalized):
            nxt = normalized[i+1]
            c_price = curr.get("Price")
            n_price = nxt.get("Price")
            c_net = curr.get("NetAmt")
            n_net = nxt.get("NetAmt")
            
            # If they share the exact same price and netamt, they MIGHT be a split item
            if c_price == n_price and c_net == n_net and c_price > 0:
                c_part = str(curr.get('Particulars', '')).strip()
                n_part = str(nxt.get('Particulars', '')).strip()
                
                c_is_tax = bool(re.search(r'\b(CGST|SGST|IGST)\b', c_part.upper()))
                n_is_tax = bool(re.search(r'\b(CGST|SGST|IGST)\b', n_part.upper()))
                
                c_has_id = bool(re.match(r'^\[.*?\]', c_part)) or c_is_tax
                n_has_id = bool(re.match(r'^\[.*?\]', n_part)) or n_is_tax
                
                is_substring = (n_part.lower() in c_part.lower()) or (c_part.lower() in n_part.lower())
                is_fragment = len(n_part) < 15 or not n_part[0].isupper() or n_part[0] in "#([{"
                
                if c_has_id and n_has_id:
                    should_merge = False
                elif c_part == n_part:
                    # Do not merge identical consecutive items (they are distinct billing entries)
                    should_merge = False
                elif is_substring or is_fragment:
                    should_merge = True
                else:
                    should_merge = False

                if should_merge:
                    curr['Particulars'] = (c_part + " " + n_part).strip()
                    i += 2
                    merged_items.append(curr)
                    continue
        
        merged_items.append(curr)
        i += 1

    # Final math correction for quantity and pricing (Triangulated Math Reconciliation)
    for item in merged_items:
        qty = item.get("Quantity", 1.0)
        price = item.get("Price", 0.0)
        net = item.get("NetAmt", 0.0)
        
        # Line-level raw text number check
        if page_text:
            line = find_matching_line(item.get("Particulars"), page_text)
            if line:
                line_no_commas = line.replace(",", "")
                net_str_dec = f"{net:.2f}"
                net_str_int = str(int(net))
                price_str_dec = f"{price:.2f}"
                price_str_int = str(int(price))
                
                net_exists = (net_str_dec in line_no_commas or net_str_int in line_no_commas) if net > 0 else True
                price_exists = (price_str_dec in line_no_commas or price_str_int in line_no_commas) if price > 0 else True
                
                # Check 1: If neither Price nor NetAmt exists in the matched raw text line,
                # but the line has exactly one currency/monetary amount (e.g. 2000.00),
                # correct hallucinated or subtotal-grabbed Price/NetAmt to the exact line currency amount.
                if not net_exists and not price_exists:
                    line_amounts = [float(x) for x in re.findall(r'\b\d+\.\d{2}\b', line_no_commas)]
                    if len(line_amounts) == 1 and line_amounts[0] > 0:
                        log.info("  [MATH RECONCILIATION] Correcting hallucinated/subtotal Price/NetAmt (%.2f) to exact line currency amount %.2f for item '%s'", net, line_amounts[0], item.get("Particulars"))
                        item["NetAmt"] = line_amounts[0]
                        item["Price"] = round(line_amounts[0] / qty, 2)
                        price = item["Price"]
                        net = item["NetAmt"]
                        net_exists = True
                        price_exists = True

                # Check 2: Check for column misalignment: Qty > 1, Price and NetAmt are different,
                # but only Price (or only NetAmt) exists in the text.
                if qty > 1.0 and abs(price - net) > 0.05:
                    if price_exists and not net_exists:
                        log.info("  [MATH RECONCILIATION] Swapping Price %.2f to NetAmt because NetAmt %.2f was not found in text line: '%s'", price, net, line.strip())
                        item["NetAmt"] = price
                        item["Price"] = round(price / qty, 2)
                        price = item["Price"]
                        net = item["NetAmt"]
                    elif net_exists and not price_exists:
                        item["Price"] = round(net / qty, 2)
                        price = item["Price"]
        
        # If Price equals NetAmt but Quantity is greater than 1,
        # it is highly likely that Price was misaligned to the total Amount column
        # and the true unit price should be NetAmt / Quantity.
        if qty > 1.0 and price == net and price > 0:
            item["Price"] = round(net / qty, 2)
            price = item["Price"]
        
        expected_net = round(qty * price, 2)
        if abs(expected_net - net) > 0.05:
            # Check if NetAmt and Price are reliable, and Qty was OCR misread
            if net > 0 and price > 0:
                calc_qty = net / price
                # If calculated quantity is close to an integer or standard decimal
                if abs(calc_qty - round(calc_qty)) < 0.05 and round(calc_qty) > 0:
                    item["Quantity"] = float(round(calc_qty))
                    item["NetAmt"] = round(item["Quantity"] * price, 2)
                    continue
            # Check if NetAmt and Qty are reliable, and Price had an OCR decimal/digit misread
            if net > 0 and qty > 0:
                calc_price = round(net / qty, 2)
                # If the calculated price differs only by a common OCR error
                if abs(calc_price - price) > 0.05:
                    item["Price"] = calc_price
                    item["NetAmt"] = round(qty * calc_price, 2)
                    continue
            # Fallback: adjust NetAmt to match Qty * Price
            item["NetAmt"] = expected_net

    return merged_items


def _extract_page(page_text: str, page_num: int, max_retries: int = 5) -> dict:
    if not page_text.strip():
        return {"page_number": page_num, "items": []}

    lines = page_text.splitlines()
    if len(lines) > 28:
        header = lines[:10]
        body = lines[10:]
        
        chunks_text = []
        current_chunk = []
        for line in body:
            current_chunk.append(line)
            # A line ends with a price/decimal if it matches standard numeric pricing pattern at the end
            has_price_end = bool(re.search(r'\b\d+(?:\.\d{2})?\s*$', line.strip()))
            
            # Split if we have reached 18 lines and it's a safe boundary, or if we hit the 28 lines limit
            if (len(current_chunk) >= 18 and has_price_end) or len(current_chunk) >= 28:
                chunks_text.append("\n".join(header + current_chunk))
                current_chunk = []
                
        if current_chunk:
            chunks_text.append("\n".join(header + current_chunk))
            
        all_chunk_items = []
        for sub_text in chunks_text:
            res = _extract_page_single(sub_text, page_num, max_retries)
            all_chunk_items.extend(res.get("items", []))
        
        seen_keys = set()
        exact_unique = []
        for it in all_chunk_items:
            if isinstance(it, dict):
                key = tuple((k, str(it[k]).strip()) for k in sorted(it.keys()) if k != "page_number")
                if key not in seen_keys:
                    seen_keys.add(key)
                    exact_unique.append(it)
        
        # Fuzzy Deduplication Pass across the page items
        import difflib
        fuzzy_unique = []
        for it in exact_unique:
            if not isinstance(it, dict):
                continue
            is_dup = False
            part_curr = str(it.get("Particulars", "")).strip()
            part_curr_clean = re.sub(r'[^a-zA-Z0-9]', '', part_curr).lower()
            p_curr = it.get("Price", 0.0)
            q_curr = it.get("Quantity", 1.0)
            net_curr = it.get("NetAmt", 0.0)
            
            for existing in fuzzy_unique:
                if isinstance(existing, dict):
                    part_ex = str(existing.get("Particulars", "")).strip()
                    part_ex_clean = re.sub(r'[^a-zA-Z0-9]', '', part_ex).lower()
                    p_ex = existing.get("Price", 0.0)
                    q_ex = existing.get("Quantity", 1.0)
                    net_ex = existing.get("NetAmt", 0.0)
                    
                    # Never fuzzy-dedup distinct tax items (e.g. CGST vs SGST are legitimate pairs)
                    tax_re = re.compile(r'\b(CGST|SGST|IGST|GST|VAT)\b', re.IGNORECASE)
                    curr_is_tax = bool(tax_re.search(part_curr))
                    ex_is_tax = bool(tax_re.search(part_ex))
                    if curr_is_tax and ex_is_tax:
                        curr_tax_type = tax_re.search(part_curr).group(1).upper()
                        ex_tax_type = tax_re.search(part_ex).group(1).upper()
                        if curr_tax_type != ex_tax_type:
                            continue  # Different tax types — keep both
                            
                    # Check for distinguishing terms (e.g. 1st visit vs subsequent visit, visit 1 vs visit 2)
                    distinguishing_pairs = [
                        ("1st", "subse"), ("first", "subse"), ("1st", "2nd"), ("first", "second"),
                        ("1st", "subsequent"), ("first", "subsequent"),
                        ("visit 1", "visit 2"), ("routine visit 1", "routine visit 2"), ("admission", "discharge"),
                        ("cgst", "sgst"), ("cgst", "igst"), ("sgst", "igst"), ("b/711", "b/719")
                    ]
                    has_distinction = False
                    for t1, t2 in distinguishing_pairs:
                        if (t1 in part_curr.lower() and t2 in part_ex.lower()) or (t2 in part_curr.lower() and t1 in part_ex.lower()):
                            has_distinction = True
                            break
                    if has_distinction:
                        continue
                        
                    # Check if they are duplicates
                    is_exact_name = (part_curr_clean == part_ex_clean and len(part_curr_clean) > 3)
                    is_fuzzy_name = (part_curr.lower() in part_ex.lower() or part_ex.lower() in part_curr.lower() or difflib.SequenceMatcher(None, part_curr.lower(), part_ex.lower()).ratio() > 0.85)
                    
                    if not is_exact_name and is_fuzzy_name:
                        s_shorter = part_curr if len(part_curr) < len(part_ex) else part_ex
                        if len(s_shorter.strip()) >= 5 and page_text.lower().count(s_shorter.lower().strip()) >= 2:
                            continue  # Both occurrences physically exist on the page — do NOT merge!
                    
                    if is_exact_name or (is_fuzzy_name and p_curr == p_ex and q_curr == q_ex and p_curr != 0 and p_curr is not None):
                        if len(part_curr) > len(part_ex):
                            existing["Particulars"] = part_curr
                        # If exact name match but different amounts, keep the one with smaller netamt or where math holds
                        if is_exact_name and net_curr != net_ex and net_curr > 0 and net_ex > 0:
                            if net_curr < net_ex:
                                existing["Price"] = p_curr
                                existing["Quantity"] = q_curr
                                existing["NetAmt"] = net_curr
                        is_dup = True
                        break
            if not is_dup:
                fuzzy_unique.append(it)
                
        return {"page_number": page_num, "items": fix_ocr_json_math(fuzzy_unique, page_text)}

    return _extract_page_single(page_text, page_num, max_retries)


def is_summary_or_category_item(name: str) -> bool:
    name_clean = re.sub(r'[^a-zA-Z0-9]', '', name).lower()
    
    # 0. Table header and section subtotal indicators
    if "total :" in name.lower() or ": total" in name.lower() or "dated -" in name.lower() or "dated ." in name.lower() or "dated :" in name.lower() or "pharm room service" in name.lower() or "billing eri" in name.lower() or "single item pharmacy" in name.lower() or "ss-pharmacy retail" in name.lower():
        return True
    if name.upper().strip() in ["RETURNED", "R;E;TU;R;N;EDI", "BATCH NUMBER", "EXPIRY DATE", "TOTAL AMOUNT", "RETURNED AMOUNT", "ITEM NUMBER", "ITEM CODE", "ITEM NAME", "BILLED QTY", "RTN QTY"]:
        return True
        
    # 1. Filter out actual final totals/footers and document headers
    total_keys = {
        "total", "subtotal", "grandtotal", "grossamount", "netamount", "amountpayable",
        "billofsupply", "netamountduers", "hospitalbillamount", "grossbillamount",
        "netbillamount", "paidamount", "netpayable", "totalclaimed", "pharmacybill",
        "pharmacysalesreturn", "billdetails", "return", "returned", "rtn", "returntotal"
    }
    if name_clean in total_keys:
        return True
        
    # 2. General category/subtotal indicators (avoiding matching "total" inside product names like "STERIPORT TOTAL")
    if name_clean == "total" or re.search(r'\b(sub\s*total|grand\s*total|gross\s*total|category\s*total|dept\s*total|group\s*total|subtotal|summary)\b', name.lower()):
        return True
    
    # 3. Known billing category/department names that are summaries, NOT individual items.
    category_summary_names = {
        "medicineippharmacy", "consumablesippharmacy", "medicinecharges",
        "consumablecharges", "consumablescharges", "pharmacydrugs",
        "pharmacysalesreturn", "particularscharges", "bedcharges",
        "pathologyinvestigation", "doctorfees", "othercharges",
        "drugsandconsumables", "nursingcharges", "procedureservicecharge",
        "administrativecharges", "dieteticsdepartment", "doctorsvisitcharge",
        "theatrecharges", "surgeonfee", "anasthetistsfees", "surgicalsupportfees",
        "cssdcharges", "implantcharges", "gstbedcharges", "sspharmacyretail",
        "billingeri", "pharmroomservice", "singleitempharmacyretail", "billofsupply",
        "netamountduers"
    }
    if name_clean in category_summary_names:
        return True
            
    # 4. Match intermediate pharmacy bills only if they explicitly contain "bill"/"bil"/"bitl"/"ein"
    # and a long alphanumeric code containing at least one digit
    if re.search(r'\b(bill|bil|bitl|ein)\b', name.lower()):
        if re.search(r'\b[a-zA-Z0-9]{5,}\b', name) and re.search(r'\d', name):
            return True
            
    return False


def _extract_page_single(page_text: str, page_num: int, max_retries: int = 5) -> dict:
    if not page_text.strip():
        return {"page_number": page_num, "items": []}
        
    headers = detect_page_headers(page_text)
    keys_instruction = "\n".join(f'- "{h}": Extract the column value for {h}.' for h in headers)
    detected_cols_str = ", ".join(headers)
    
    sample_item = {}
    for h in headers:
        if h in ["SI#", "Sl#"]:
            sample_item[h] = "1"
        elif h in ["Qty", "Quantity"]:
            sample_item[h] = 1.0
        elif h in ["Rate", "Price"]:
            sample_item[h] = 150.0
        elif h in ["Gross Amount", "NetAmt"]:
            sample_item[h] = 150.0
        elif h == "Discount":
            sample_item[h] = 0.0
        elif h == "Date":
            sample_item[h] = "13/06/2026"
        elif h == "Cpt Code":
            sample_item[h] = ""
        else:
            sample_item[h] = "Item Name"
            
    sample_str = json.dumps(sample_item)
    
    ocr_instructions = ""
    user_prompt = f"""Extract every single financial/billing line item from the invoice/receipt text below.

DETECTED TABLE COLUMNS FOR THIS PAGE: [{detected_cols_str}]
For each line item, extract the data and map it EXACTLY to the following JSON keys:
{keys_instruction}

OUTPUT FORMAT — respond with ONLY this JSON structure, nothing else:
{{
  "page_number": {page_num},
  "items": [
    {sample_str}
  ]
}}

RULES:
- Only extract FINANCIAL billing line items. If a page or a list contains NO financial prices (e.g. a medical report, clinical notes, patient details), IGNORE IT completely and return an empty list for "items".
- Do NOT extract patient metadata (Name, Age, Sex, Address, Dates, etc.) as line items.
- Extract the complete item description exactly as it appears. Do not truncate.
- All numeric fields must be plain numbers (no $ signs, no commas).
- Do NOT collapse or deduplicate repeating items. If the document has multiple identical lines, you must output each one as a separate distinct item in the JSON list. Never skip or aggregate repeating charges.
- Do NOT extract category subtotal lines or summary rows (such as lines ending with 'Sub Total: ...'). Extract ONLY individual itemized charges.
- Do NOT extract Services Accounting Codes (SAC) or HSN codes (typically formatted as "SAC:XXXXXX" or "HSN:XXXXXX" or as 6-digit integers starting with 99) as prices or line items. They are tax classifications, NOT financial amounts.
- You MUST use the EXACT numeric values printed in the document. Never change, guess, or synthesize numbers. If the text says 4,620.00, you must output 4620.00. Outputting a slightly different number is a critical error and will cause the item to be discarded.
{ocr_instructions}

DO NOT write any sentence or paragraph. START your response with the {{ character.

--- INVOICE TEXT ---
{page_text}
--- END ---
"""

    messages = [
        {"role": "system", "content": SYSTEM_PROMPT % page_num},
        {"role": "user",   "content": user_prompt},
    ]
    
    label = f"page {page_num}"
    for attempt in range(1, max_retries + 1):
        reply = _call_llm(messages, f"{label} attempt {attempt}")
        if not reply:
            log.warning(f"  [LLM] Ollama returned no reply on attempt {attempt}. Retrying...")
            import time
            time.sleep(5)
            continue
            
        result = _parse_reply(reply, page_num)
        if result is not None:
            items = result.get("items") or []
            items = fix_ocr_json_math(items, page_text)
            
            # Post-processing: Filter out category/summary rows and duplicates
            filtered_items = []
            for item in items:
                if not isinstance(item, dict):
                    continue
                part = item.get("Particulars", "")
                if is_summary_or_category_item(part):
                    log.info("  [CATEGORY FILTER] Filtering out summary/category item: %s", part)
                    continue
                
                # Failsafe check to filter out items without any letters (e.g. subtotals like "2260.00" or empty)
                if not re.search(r'[a-zA-Z]', part):
                    log.info("  [LETTER FILTER] Filtering out item with no alphabetic characters: %s", part)
                    continue
                
                # Failsafe check to filter out leftover header names and SAC/HSN code patterns
                part_upper = part.upper().strip()
                if part_upper in ["ORDER ITEM", "SL# DESCRIPTION DATE QTY RATE GROSS AMOUNT DISCOUNT", "PARTICULARS QTY PRICE NETAMT"]:
                    continue
                if "SAC:" in part_upper or "HSN:" in part_upper:
                    continue
                
                # Filter out general SAC codes starting with 99 (6-digit integer check)
                price_val = item.get("Price")
                net_val = item.get("NetAmt")
                is_sac = False
                for val in [price_val, net_val]:
                    if val is not None and val != 0 and val != 0.0:
                        try:
                            f_val = float(str(val).replace(",", "").strip())
                            if f_val.is_integer():
                                i_val = int(f_val)
                                if 990000 <= i_val <= 999999:
                                    is_sac = True
                                    break
                        except ValueError:
                            pass
                if is_sac:
                    log.info("  [SAC FILTER] Filtering out item '%s' because its price/netamt (%s/%s) matches an SAC code pattern.", part, price_val, net_val)
                    continue
                vals = str(item.values()).lower()
                if "fake_" in vals or "xyz123" in vals or "999.99" in vals:
                    continue
                    
                # Filter specific LLM hallucinations that occur on empty/summary pages
                is_hallucinated = False
                desc_full = ""
                for k, v in item.items():
                    v_str = str(v).lower().strip()
                    if v_str in ["medicine", "ip charges", "materials", "implant charges", "diet",
                                 "medicine-ip pharmacy", "consumables-ip pharmacy", "medicine charges"]:
                        is_hallucinated = True
                    if isinstance(v, str) and not v.replace(".","",1).isdigit():
                        desc_full += v.lower()
                clean_desc = re.sub(r'[^a-z0-9]', '', desc_full)
                if is_hallucinated or clean_desc in ["medicine", "ipcharges", "materials", "implantcharges", "diet",
                        "medicineippharmacy", "consumablesippharmacy", "medicinecharges",
                        "billinger1", "pharmroomservice", "singleitempharmacyretail",
                        "total", "subtotal", "grossamount", "netamount", "mrdno", "patientname"]:
                    continue
                # Filter out standalone reference bill numbers only on summary page 1
                if page_num == 1 and bool(re.match(r'^bill\d+$', clean_desc)):
                    continue
                        
                if "id" in item and len(item) <= 2 and str(item.get("text")).lower() == "null":
                    continue
                if "mrd number" in vals or "patient name" in vals or "company name" in vals:
                    continue
                
                # Filter out medication dosages hallucinated as financial items
                if re.search(r'\b\d-\d-\d\b', vals) or " sos " in f" {vals} " or "to continue" in vals or "if sbp" in vals or "mmhg" in vals:
                    continue
                    
                # Filter out variations of Pharm Room Service OCR garbage
                if "room" in vals and ("pharm" in vals or "phatm" in vals or "serv" in vals or "sery" in vals):
                    continue
                
                # Drop purely empty/null hallucinated rows
                is_empty = True
                for k, v in item.items():
                    # Check if value is meaningful (not null, not 0.0, not empty string)
                    if v is not None and v != "" and v != 0 and v != 0.0 and str(v).lower() != "null":
                        is_empty = False
                        break
                if is_empty:
                    continue
                    
                # Drop items that have no financial value/price attached
                has_financial = False
                for k, v in item.items():
                    if v is None or str(v).strip() == "" or str(v).lower() == "null" or str(v) == "None":
                        continue
                        
                    k_lower = k.lower()
                    # If it's a known financial column and it has a value != 0, it's financial
                    if "amount" in k_lower or "charge" in k_lower or "price" in k_lower or "rate" in k_lower or "netamt" in k_lower or "total" in k_lower or "cost" in k_lower:
                        try:
                            if float(str(v).replace(",", "")) != 0:
                                has_financial = True
                                break
                        except ValueError:
                            pass
                        
                    # If any value is a standalone monetary-looking number != 0, consider it financial
                    try:
                        val_str = str(v).replace(",", "")
                        if float(val_str) != 0 and k_lower not in ["quantity", "qty", "sl#", "sr no", "sr.no", "sn", "no", "item code", "code", "batch", "sac", "hsn"]:
                            has_financial = True
                            break
                    except ValueError:
                        pass
                
                if not has_financial:
                    continue
                    
                # ULTIMATE HALLUCINATION FILTER:
                # Ensure at least one significant extracted value physically exists in the raw OCR text!
                is_real = False
                raw_text_lower = page_text.lower()
                raw_text_no_commas = page_text.replace(",", "")
                for v in item.values():
                    if v is None or str(v).strip() == "" or str(v).lower() == "null" or str(v) == "None":
                        continue
                    v_str = str(v).replace(",", "").strip()
                    
                    # 1. Check if it's a numeric value > 0 that exists in the text
                    try:
                        f_val = float(v_str)
                        if f_val > 0 and f_val not in [1.0, 2.0]:
                            if str(f_val) in raw_text_no_commas or str(int(f_val)) in raw_text_no_commas or v_str in raw_text_no_commas:
                                is_real = True
                                break
                            if f"{f_val:.2f}" in raw_text_no_commas or f"{int(f_val)}.00" in raw_text_no_commas or f"{f_val:.1f}" in raw_text_no_commas:
                                is_real = True
                                break
                    except ValueError:
                        pass
                        
                    # 2. Check if it's a non-numeric string (e.g. Description) that exists in the text
                    if isinstance(v, str) and len(v) > 4:
                        v_lower = v.lower()
                        # Ignore generic LLM hallucinated words that might accidentally be on the page
                        if v_lower not in ["medicine", "medication", "consultation fee", "ip charges", "room charge", "pharmacy", "amount", "total", "total amount", "gross amount", "net amount"]:
                            # Try exact direct match first
                            if v_lower[:15] in raw_text_lower or v_lower in raw_text_lower:
                                is_real = True
                                break
                            
                            # Fallback to sliding-window fuzzy matching to handle OCR word splits/merges
                            import difflib
                            words = raw_text_lower.split()
                            v_words = v_lower.split()
                            v_len = len(v_words)
                            if v_len > 0:
                                for idx in range(max(1, len(words) - v_len + 1)):
                                    window = " ".join(words[idx : idx + v_len])
                                    ratio = difflib.SequenceMatcher(None, v_lower, window).ratio()
                                    if ratio >= 0.75:  # 75% similarity threshold
                                        is_real = True
                                        break
                                if is_real:
                                    break
                
                if not is_real:
                    continue
                    
                # Ground Truth Price/NetAmt Verification:
                # Ensure the extracted price or net amount (if > 2.0) physically exists in the text.
                has_valid_price = False
                price_val = item.get("Price")
                net_val = item.get("NetAmt")
                
                has_any_val = False
                for v in [price_val, net_val]:
                    if v is not None and v != 0 and v != 0.0 and str(v).lower() != "null":
                        has_any_val = True
                        try:
                            f_val = float(str(v).replace(",", ""))
                            if f_val <= 2.0:
                                has_valid_price = True
                                break
                            val_str = f"{f_val:.2f}"
                            val_str_no_dec = str(int(f_val))
                            val_str_one_dec = f"{f_val:.1f}"
                            if val_str in raw_text_no_commas or val_str_no_dec in raw_text_no_commas or val_str_one_dec in raw_text_no_commas:
                                has_valid_price = True
                                break
                        except ValueError:
                            pass
                
                if has_any_val and not has_valid_price:
                    log.info(f"  [HALLUCINATION FILTER] Discarding item '{item.get('Particulars')}' because its price/netamt ({price_val}/{net_val}) was not found in the raw text.")
                    continue
                    
                filtered_items.append(item)
            
            result["items"] = filtered_items
            log.info("  [LLM] %s → %d item(s) on attempt %d.", label, len(filtered_items), attempt)
            return result
            
        log.warning("  [LLM] %s attempt %d returned non-JSON — retrying…", label, attempt)
        messages = [
            {"role": "system",    "content": SYSTEM_PROMPT % page_num},
            {"role": "user",      "content": USER_PROMPT_TEMPLATE.format(
                page_num=page_num, raw_text=page_text, ocr_instructions=ocr_instructions)},
            {"role": "assistant", "content": reply},
            {"role": "user",      "content": RETRY_PROMPT_TEMPLATE.format(
                page_num=page_num, raw_text=page_text[:4000])},
        ]

    log.error("  [LLM] %s — all %d attempts failed.", label, max_retries)
    return {"page_number": page_num, "items": []}


def extract_deterministic_table_rows_from_pdf(file_path: Path) -> list[dict]:
    import pdfplumber
    try:
        pdf = pdfplumber.open(file_path)
    except Exception as e:
        log.warning("  [DETERMINISTIC EXTRACTOR] Could not open %s with pdfplumber: %s", file_path, e)
        return []
        
    pages_results = []
    cat_headers = [
        "ADMINISTRATIVE CHARGES", "BED CHARGES", "DIETETICS DEPARTMENT", 
        "DOCTORS VISIT CHARGE", "LABORATORY", "NURSING CHARGES", "PROCEDURE/SERVICE CHARGE", 
        "RADIOLOGY", "CONSUMABLES-IP PHARMACY", "MEDICINE-IP PHARMACY", "MATERIALS", 
        "IMPLANT CHARGES", "THEATRE CHARGES", "SURGEON FEE", "ANASTHETISTS FEES", 
        "SURGICAL SUPPORT FEES", "CSSD CHARGES", "MISCELLANEOUS", "PHARMACY DRUGS",
        "Category Total", "Gross Total", "Net Amount", "Amount to be claimed", "Claimed Amount", 
        "Remarks", "Prepared By", "IP COUNTER"
    ]
    
    for p_idx, p in enumerate(pdf.pages):
        words = p.extract_words()
        if not words:
            pages_results.append({"page_number": p_idx + 1, "items": [], "taxes": []})
            continue
            
        sorted_words = sorted(words, key=lambda w: (round(w['top'] / 5.0) * 5.0, w['x0']))
        rows = {}
        for w in sorted_words:
            row_y = round(w['top'] / 5.0) * 5.0
            rows.setdefault(row_y, []).append(w['text'])
            
        items = []
        current_item = None
        
        for row_y, l in sorted(rows.items()):
            text = " ".join(l).strip()
            if not text:
                continue
                
            if any(h.upper() in text.upper() for h in ["PAN No", "Mobile No", "MRD No", "IP No", "Claim No", "Page ", "IP BILL BREAKUP", "Patient Name", "Admission Date", "Discharge Date", "Customer", "Bill No", "Bill Type", "Bill Date", "Doctor", "Department", "Bed No", "Sl# Description", "Cpt Code"]):
                continue
            if any(kw in text.lower() for kw in ["net amount", "gross total", "total 34,", "discount ", "rupees", "thousand", "hundred", "claimed amount", "remarks:", "prepared by", "counter", "amount to be claimed", "tax)", "only"]):
                continue
                
            m_sl = re.match(r"^([1-9][0-9]{0,2})\s+(.+)$", text)
            if m_sl:
                sl_num = int(m_sl.group(1))
                body = m_sl.group(2).strip()
                
                m_end = re.search(r"\b(\d{2}[/-]\d{2}[/-]\d{4})\s+([\d,.]+)\s+([\d,.]+)\s+([\d,.]+)\s+([\d,.]+)$", body)
                if not m_end:
                    m_end = re.search(r"\b(\d{2}[/-]\d{2}[/-]\d{4})\s+([\d,.]+)\s+([\d,.]+)\s+([\d,.]+)$", body)
                if not m_end:
                    m_end = re.search(r"\s+([\d,.]+)\s+([\d,.]+)\s+([\d,.]+)$", body)
                    
                if m_end:
                    groups = m_end.groups()
                    if len(groups) == 5:
                        date_val, qty_val, rate_val, gross_val, disc_val = groups
                    elif len(groups) == 4:
                        if re.match(r"\d{2}[/-]\d{2}[/-]\d{4}", groups[0]):
                            date_val, qty_val, rate_val, gross_val = groups
                            disc_val = "0.00"
                        else:
                            date_val = ""
                            qty_val, rate_val, gross_val, disc_val = groups
                    else:
                        date_val = ""
                        rate_val, gross_val, disc_val = groups
                        qty_val = "1"
                        
                    desc = body[:m_end.start()].strip()
                    
                    def clean_num(s):
                        try:
                            return float(s.replace(",", "").strip())
                        except:
                            return 0.0
                            
                    current_item = {
                        "Sl#": str(sl_num),
                        "Description": desc,
                        "Cpt Code": "",
                        "Date": date_val,
                        "Qty": clean_num(qty_val),
                        "Quantity": clean_num(qty_val),
                        "Rate": clean_num(rate_val),
                        "Price": clean_num(rate_val),
                        "Gross Amount": clean_num(gross_val),
                        "NetAmt": clean_num(gross_val),
                        "Discount": clean_num(disc_val),
                        "item_type": "item",
                        "Particulars": desc,
                    }
                    items.append(current_item)
                    continue
                    
            if any(h.upper() == text.upper() or text.upper().startswith(h.upper() + " ") for h in cat_headers) or text.upper() == "ADMISSION":
                continue
                
            if current_item and not any(c.isdigit() for c in text[:2]):
                if not any(kw in text.lower() for kw in ["total", "discount", "observation in op/casualty"]):
                    current_item["Description"] += " " + text
                    current_item["Description"] = current_item["Description"].strip()
                    current_item["Particulars"] = current_item["Description"]
                
        pages_results.append({"page_number": p_idx + 1, "items": items, "taxes": []})
        
    return pages_results


def extract_deterministic_table_rows_from_text(pages_text: list[str]) -> list[dict]:
    pages_results = []
    cat_headers = [
        "ADMINISTRATIVE CHARGES", "BED CHARGES", "DIETETICS DEPARTMENT", 
        "DOCTORS VISIT CHARGE", "LABORATORY", "NURSING CHARGES", "PROCEDURE/SERVICE CHARGE", 
        "RADIOLOGY", "CONSUMABLES-IP PHARMACY", "MEDICINE-IP PHARMACY", "MATERIALS", 
        "IMPLANT CHARGES", "THEATRE CHARGES", "SURGEON FEE", "ANASTHETISTS FEES", 
        "SURGICAL SUPPORT FEES", "CSSD CHARGES", "MISCELLANEOUS", "PHARMACY DRUGS",
        "Category Total", "Gross Total", "Net Amount", "Amount to be claimed", "Claimed Amount", 
        "Remarks", "Prepared By", "IP COUNTER"
    ]
    
    for p_idx, p_text in enumerate(pages_text):
        if not p_text.strip():
            pages_results.append({"page_number": p_idx + 1, "items": [], "taxes": []})
            continue
            
        lines = p_text.split('\n')
        items = []
        current_item = None
        
        for text in lines:
            text = text.strip()
            if not text:
                continue
                
            if any(h.upper() in text.upper() for h in ["PAN No", "Mobile No", "MRD No", "IP No", "Claim No", "Page ", "IP BILL BREAKUP", "Patient Name", "Admission Date", "Discharge Date", "Customer", "Bill No", "Bill Type", "Bill Date", "Doctor", "Department", "Bed No", "Sl# Description", "Cpt Code"]):
                continue
            if any(kw in text.lower() for kw in ["net amount", "gross total", "total 34,", "discount ", "rupees", "thousand", "hundred", "claimed amount", "remarks:", "prepared by", "counter", "amount to be claimed", "tax)", "only"]):
                continue
                
            m_sl = re.match(r"^([1-9][0-9]{0,2})\s+(.+)$", text)
            if m_sl:
                sl_num = int(m_sl.group(1))
                body = m_sl.group(2).strip()
                
                m_end = re.search(r"\b(\d{2}[/-]\d{2}[/-]\d{4})\s+([\d,.]+)\s+([\d,.]+)\s+([\d,.]+)\s+([\d,.]+)$", body)
                if not m_end:
                    m_end = re.search(r"\b(\d{2}[/-]\d{2}[/-]\d{4})\s+([\d,.]+)\s+([\d,.]+)\s+([\d,.]+)$", body)
                if not m_end:
                    m_end = re.search(r"\s+([\d,.]+)\s+([\d,.]+)\s+([\d,.]+)$", body)
                    
                if m_end:
                    groups = m_end.groups()
                    if len(groups) == 5:
                        date_val, qty_val, rate_val, gross_val, disc_val = groups
                    elif len(groups) == 4:
                        if re.match(r"\d{2}[/-]\d{2}[/-]\d{4}", groups[0]):
                            date_val, qty_val, rate_val, gross_val = groups
                            disc_val = "0.00"
                        else:
                            date_val = ""
                            qty_val, rate_val, gross_val, disc_val = groups
                    else:
                        date_val = ""
                        rate_val, gross_val, disc_val = groups
                        qty_val = "1"
                        
                    desc = body[:m_end.start()].strip()
                    
                    def clean_num(s):
                        try:
                            return float(s.replace(",", "").strip())
                        except:
                            return 0.0
                            
                    current_item = {
                        "Sl#": str(sl_num),
                        "Description": desc,
                        "Cpt Code": "",
                        "Date": date_val,
                        "Qty": clean_num(qty_val),
                        "Quantity": clean_num(qty_val),
                        "Rate": clean_num(rate_val),
                        "Price": clean_num(rate_val),
                        "Gross Amount": clean_num(gross_val),
                        "NetAmt": clean_num(gross_val),
                        "Discount": clean_num(disc_val),
                        "item_type": "item",
                        "Particulars": desc,
                    }
                    items.append(current_item)
                    continue
                    
            if any(h.upper() == text.upper() or text.upper().startswith(h.upper() + " ") for h in cat_headers) or text.upper() == "ADMISSION":
                continue
                
            if current_item and not any(c.isdigit() for c in text[:2]):
                if not any(kw in text.lower() for kw in ["total", "discount", "observation in op/casualty"]):
                    current_item["Description"] += " " + text
                    current_item["Description"] = current_item["Description"].strip()
                    current_item["Particulars"] = current_item["Description"]
                
        pages_results.append({"page_number": p_idx + 1, "items": items, "taxes": []})
        
    return pages_results


def extract_with_llm(pages_text: list[str], filename: str, file_path: Optional[Path] = None) -> Optional[dict]:
    from concurrent.futures import ThreadPoolExecutor, as_completed

    total_pages = len(pages_text)
    log.info("  [LLM] Document has %d page(s) — processing in parallel.", total_pages)

    pages_results = []
    if file_path and file_path.exists():
        det_pages = extract_deterministic_table_rows_from_pdf(file_path)
        det_total_items = sum(len(p.get("items", [])) for p in det_pages)
        if det_total_items >= 5:
            log.info("  [DETERMINISTIC EXTRACTOR] Found %d structured table rows via spatial PDF alignment in %s! Bypassing LLM.", det_total_items, filename)
            pages_results = det_pages

    if not pages_results:
        det_text_pages = extract_deterministic_table_rows_from_text(pages_text)
        det_text_total = sum(len(p.get("items", [])) for p in det_text_pages)
        if det_text_total >= 5:
            log.info("  [DETERMINISTIC EXTRACTOR] Found %d structured table rows via text regex alignment in %s! Bypassing LLM.", det_text_total, filename)
            pages_results = det_text_pages

    if not pages_results:
        results_by_index: dict[int, dict] = {}

        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
            future_to_idx = {
                pool.submit(_extract_page, text, i+1): i
                for i, text in enumerate(pages_text)
            }
            for future in as_completed(future_to_idx):
                idx = future_to_idx[future]
                try:
                    results_by_index[idx] = future.result()
                except Exception as exc:
                    log.error("  [LLM] page %d raised an exception: %s", idx + 1, exc)
                    results_by_index[idx] = {"page_number": idx+1, "items": []}

        # Reassemble in original document order
        for idx in range(total_pages):
            res = results_by_index.get(idx, {"page_number": idx+1, "items": []})
            page_num = idx + 1
            raw_items = res.get("items", [])
            page_items = []
            page_taxes = []

            for item in raw_items:
                if not isinstance(item, dict):
                    continue
                new_item = normalize_item(item)
                if new_item.get("Particulars"):
                    if new_item.get("item_type") == "tax":
                        page_taxes.append(new_item)
                    else:
                        page_items.append(new_item)

            clean_res = {
                "page_number": page_num,
                "items": page_items,
                "taxes": page_taxes
            }
            pages_results.append(clean_res)

    # ── Summary Page Deduplication Check ─────────────────────────────────────
    def is_summary_page(items):
        if not items:
            return False
            
        category_names = {
            "bedcharges", "pathologyinvestigation", "doctorfees", "procedures",
            "othercharges", "drugsandconsumables", "pharmacysalesreturn", "particularscharges",
            "administrativecharges", "admission", "admissioncharges", "dieteticsdepartment",
            "diet", "doctorsvisitcharge", "laboratory", "nursingcharges",
            "procedureservicecharge", "radiology", "pharmacydrugs", "materials",
            "implantcharges", "theatrecharges", "surgeonfee", "anasthetistsfees",
            "surgicalsupportfees", "cssdcharges", "miscellaneous",
            "billofsupply", "netamountduers", "hospitalbillamount", "grossbillamount",
            "netbillamount", "paidamount", "billdetails", "pharmacybill"
        }
        
        count = 0
        for item in items:
            part_clean = re.sub(r'[^a-zA-Z0-9]', '', str(item.get("Particulars", ""))).lower()
            if part_clean in category_names:
                count += 1
                
        return (count / len(items)) >= 0.50

    def remove_summary_pages(pages_list):
        # 1. Identify and drop pages that are high-level category summary bills
        non_summary_pages = []
        for p in pages_list:
            p_items = p.get("items", [])
            if is_summary_page(p_items):
                log.info("  [DEDUPLICATION] Page %d is identified as a summary page. Dropping Page %d.", p["page_number"], p["page_number"])
            else:
                non_summary_pages.append(p)
                
        pages_list = non_summary_pages
        page_sums = []
        for p in pages_list:
            p_sum = 0.0
            for item in p.get("items", []):
                p_sum += item.get("NetAmt", 0.0)
            page_sums.append(p_sum)
            
        n_pages = len(pages_list)
        if n_pages <= 1:
            return pages_list
            
        # 2. Page-to-Page comparison: if Page A matches Page B sum, and Page A has fewer items, drop Page A
        for i in range(n_pages):
            for j in range(n_pages):
                if i != j:
                    s_i = page_sums[i]
                    s_j = page_sums[j]
                    if s_i > 0 and abs(s_i - s_j) < max(10.0, s_i * 0.05):
                        num_i = len(pages_list[i].get("items", []))
                        num_j = len(pages_list[j].get("items", []))
                        if num_i < num_j:
                            log.info("  [DEDUPLICATION] Page %d is identified as a summary of Page %d (sum %.2f matches sum %.2f). Dropping Page %d.", pages_list[i]["page_number"], pages_list[j]["page_number"], s_i, s_j, pages_list[i]["page_number"])
                            pages_list_copy = [p for k, p in enumerate(pages_list) if k != i]
                            return remove_summary_pages(pages_list_copy)
            
        # 3. Page-to-Others comparison: if Page K matches sum of all other pages combined, and Page K has fewer items, drop Page K
        for k in range(n_pages):
            s_k = page_sums[k]
            if s_k <= 0:
                continue
            other_sum = sum(page_sums[j] for j in range(n_pages) if j != k)
            other_items_count = sum(len(pages_list[j].get("items", [])) for j in range(n_pages) if j != k)
            num_items_k = len(pages_list[k].get("items", []))
            
            # If s_k matches sum of other pages, and has fewer items, drop page k (summary page)
            if num_items_k < other_items_count and abs(s_k - other_sum) < max(10.0, s_k * 0.05):
                log.info("  [DEDUPLICATION] Page %d is identified as a summary page (sum %.2f matches sum of other pages %.2f). Dropping Page %d to prevent double-counting.", pages_list[k]["page_number"], s_k, other_sum, pages_list[k]["page_number"])
                pages_list_copy = [p for i, p in enumerate(pages_list) if i != k]
                return remove_summary_pages(pages_list_copy)
        return pages_list

    pages_results = remove_summary_pages(pages_results)
    
    # Recalculate total items (including taxes)
    total_items = sum(len(p.get("items", [])) + len(p.get("taxes", [])) for p in pages_results)

    if total_items == 0 and total_pages > 0:
        log.warning("  [LLM] No items extracted from any pages.")

    log.info("  [LLM] Total items extracted across all pages: %d", total_items)

    # Automated Math Self-Correction / Reconciliation Check
    try:
        extracted_sum = 0.0
        for p in pages_results:
            for item in p.get("items", []):
                extracted_sum += item.get("NetAmt", 0.0)
        extracted_sum = round(extracted_sum, 2)

        doc_full_text = "\n".join(pages_text)
        
        # Parse for declared invoice total elements
        subtotal_matches = re.findall(
            r"(?:sub\s*total|gross\s+amount|total\s+charges)[\s:=-]+(?:rs\.?|inr)?\s*([\d,]+\.\d{2})",
            doc_full_text,
            re.IGNORECASE,
        )
        tax_matches = re.findall(
            r"(?:tax|cgst|sgst|igst|vat|gst)[\s:=-]+(?:rs\.?|inr)?\s*([\d,]+\.\d{2})",
            doc_full_text,
            re.IGNORECASE,
        )
        discount_matches = re.findall(
            r"(?:discount|concession|rebate)[\s:=-]+(?:rs\.?|inr)?\s*([\d,]+\.\d{2})",
            doc_full_text,
            re.IGNORECASE,
        )
        total_matches = re.findall(
            r"(?:grand\s+total|gross\s+total|net\s+payable|total\s+claimed|total\s+amount|amount\s+payable|net\s+amount)[\s:=-]+(?:rs\.?|inr)?\s*([\d,]+\.\d{2})",
            doc_full_text,
            re.IGNORECASE,
        )

        subtotal_val = max([float(m.replace(",", "")) for m in subtotal_matches]) if subtotal_matches else 0.0
        tax_val = sum([float(m.replace(",", "")) for m in tax_matches]) if tax_matches else 0.0
        discount_val = max([float(m.replace(",", "")) for m in discount_matches]) if discount_matches else 0.0
        target_total = max([float(m.replace(",", "")) for m in total_matches]) if total_matches else 0.0

        if target_total == 0.0 and subtotal_val > 0.0:
            target_total = subtotal_val + tax_val - discount_val

        if target_total > 0.0:
            diff = round(target_total - extracted_sum, 2)
            
            # If extracted sum is larger than the target total, check if we included tax/subtotal lines by mistake
            if diff < -1.0:
                log.info("  [RECONCILIATION] Extracted sum (%.2f) exceeds target total (%.2f). Checking for duplicate totals/taxes...", extracted_sum, target_total)
                raw_subtotals = [float(m.group(1)) for m in re.finditer(r'Sub\s*Total\s*:\s*(\d+(?:\.\d{2})?)', doc_full_text, re.IGNORECASE)]
                for p in pages_results:
                    filtered = []
                    for item in p.get("items", []):
                        part = item.get("Particulars", "").upper().strip()
                        net_val = round(item.get("NetAmt", 0.0), 2)
                        
                        # Filter out totals, taxes, and subtotal rows
                        is_total_line = any(k in part for k in ["SUBTOTAL", "SUB TOTAL", "GRAND TOTAL", "NET PAYABLE", "TOTAL AMOUNT"])
                        is_subtotal_match = any(abs(net_val - st_val) < 0.1 for st_val in raw_subtotals) and part in ["ORDER ITEM", "SUB TOTAL", "SUBTOTAL"]
                        is_tax_line = any(k in part for k in ["CGST", "SGST", "IGST", "TAX"]) and abs(net_val - tax_val) < 2.0
                        
                        if is_total_line or is_subtotal_match or is_tax_line:
                            log.info("  [RECONCILIATION] Filtering out non-item line: %s (Amt: %.2f)", item.get("Particulars"), item.get("NetAmt"))
                            extracted_sum -= item.get("NetAmt", 0.0)
                            continue
                        filtered.append(item)
                    p["items"] = filtered
                extracted_sum = round(extracted_sum, 2)
                diff = round(target_total - extracted_sum, 2)

            # Three-Way Accounting Equation Check: check if extracted base charges + taxes or - discount balance to target_total
            if abs(diff) > 1.0:
                if abs((extracted_sum + tax_val - discount_val) - target_total) <= 2.0 or abs((extracted_sum + tax_val) - target_total) <= 2.0 or abs((extracted_sum - discount_val) - target_total) <= 2.0:
                    log.info("  [RECONCILIATION] Three-Way Accounting Check verified balanced! Extracted base sum (%.2f) with taxes/discounts matches target total (%.2f).", extracted_sum, target_total)
                else:
                    log.warning("  [RECONCILIATION] Extracted sum (%.2f) != Invoice Total (%.2f). Discrepancy: %+.2f", extracted_sum, target_total, diff)
                    if diff > 1.0 and len(pages_results) > 0:
                        log.info("  [RECONCILIATION] Triggering AI Self-Correction loop to recover missing $%.2f (Escalating to heavy accuracy model %s)...", diff, OLLAMA_HEAVY_MODEL)
                        rec_prompt = [
                            {"role": "system", "content": "You are a hospital billing audit AI."},
                            {"role": "user", "content": f"Invoice text:\n{doc_full_text[:5000]}\n\nExtracted items total {extracted_sum}, but document footer total is {target_total}. Find the missing item costing approximately {diff}. Return ONLY JSON format: {{\"items\": [{{\"Particulars\": \"missing item\", \"Quantity\": 1, \"Price\": {diff}, \"NetAmt\": {diff}}}]}}"}
                        ]
                        rec_reply = _call_llm(rec_prompt, "self-correction recovery", model_override=OLLAMA_HEAVY_MODEL)
                        if rec_reply:
                            rec_clean = _parse_reply(rec_reply, 99)
                            if rec_clean and rec_clean.get("items"):
                                recovered = rec_clean["items"]
                                log.info("  [RECONCILIATION] Self-Correction recovered %d missing item(s)!", len(recovered))
                                # Add recovered items to the last page after normalizing them
                                norm_recovered = [normalize_item(it) for it in recovered]
                                pages_results[-1]["items"].extend(norm_recovered)
                                total_items += len(norm_recovered)
                                # Update extracted_sum and diff after recovery
                                for r_item in norm_recovered:
                                    extracted_sum += r_item.get("NetAmt", 0.0)
                                extracted_sum = round(extracted_sum, 2)
                                diff = round(target_total - extracted_sum, 2)
                                if abs(diff) <= 2.0:
                                    log.info("  [RECONCILIATION] Balanced after self-correction! Extracted sum (%.2f) matches Invoice Total (%.2f).", extracted_sum, target_total)
            else:
                log.info("  [RECONCILIATION] 100%% Balanced! Extracted sum (%.2f) matches Invoice Total (%.2f).", extracted_sum, target_total)
    except Exception as e:
        log.warning("  [RECONCILIATION] Error during verification loop: %s", e)

    # Denormalize items back to original keys matching the page columns
    cat_prefixes = [
        r"ADMINISTRATIVE CHARGES", r"ADMISSION", r"BED CHARGES", r"DIETETICS DEPARTMENT",
        r"DOCTORS VISIT CHARGE", r"LABORATORY", r"NURSING CHARGES", r"PROCEDURE/SERVICE CHARGE",
        r"RADIOLOGY", r"CONSUMABLES-IP PHARMACY", r"MEDICINE-IP PHARMACY", r"MATERIALS",
        r"IMPLANT CHARGES", r"DIET", r"THEATRE CHARGES", r"SURGEON FEE", r"ANASTHETISTS FEES",
        r"SURGICAL SUPPORT FEES", r"CSSD CHARGES", r"MISCELLANEOUS", r"PHARMACY DRUGS",
        r"PARTICULARS CHARGES"
    ]
    last_sl_num = 0
    last_seen_date = ""

    for p in pages_results:
        page_num = p["page_number"]
        p_text = pages_text[page_num - 1]
        headers = detect_page_headers(p_text)
        
        denorm_items = []
        for item in p.get("items", []):
            denorm = denormalize_item(item, headers)
            
            # Find which key in denorm represents serial number
            sl_key = next((k for k in denorm if k.lower() in ["sl#", "si#"]), None)
            if not sl_key and any(h.lower() in ["sl#", "si#"] for h in headers):
                sl_key = next((h for h in headers if h.lower() in ["sl#", "si#"]), "Sl#")
                
            if sl_key:
                val = str(denorm.get(sl_key, "")).strip()
                if val.isdigit() and int(val) > last_sl_num:
                    last_sl_num = int(val)
                elif last_sl_num > 0:
                    last_sl_num += 1
                    denorm[sl_key] = str(last_sl_num)
                elif not val:
                    # Check if page text starts with Sl numbers (like 1, 2, 3)
                    if any(re.match(r"^\s*([1-9][0-9]{0,2})\s+", line) for line in p_text.splitlines()[:30]):
                        last_sl_num = 1
                        denorm[sl_key] = str(last_sl_num)
                    
            # Clean Description / Particulars
            for desc_key in ["Description", "Particulars"]:
                if desc_key in denorm:
                    desc = str(denorm[desc_key]).strip()
                    for cat in cat_prefixes:
                        if re.match(r"^" + cat + r"\b", desc, flags=re.IGNORECASE):
                            desc = re.sub(r"^" + cat + r"\s*", "", desc, flags=re.IGNORECASE).strip()
                            break
                    m_sl = re.match(r"^([0-9]{1,4})\s+(.*)$", desc)
                    if m_sl and 0 < int(m_sl.group(1)) <= 500:
                        desc = m_sl.group(2).strip()
                        if sl_key and not denorm.get(sl_key):
                            denorm[sl_key] = m_sl.group(1)
                            if m_sl.group(1).isdigit() and int(m_sl.group(1)) > last_sl_num:
                                last_sl_num = int(m_sl.group(1))
                    # Remove trailing Sl number if duplicated at end of description
                    if sl_key and denorm.get(sl_key):
                        m_end = re.search(r"\s+([0-9]{1,4})$", desc)
                        if m_end and m_end.group(1) == str(denorm.get(sl_key)):
                            desc = re.sub(r"\s+" + m_end.group(1) + r"$", "", desc).strip()
                    denorm[desc_key] = desc

            # Auto-populate Date from exact line in raw text or fallback to last seen date
            if "Date" in denorm or any(h.lower() == "date" for h in headers):
                date_key = "Date" if "Date" in denorm else next((h for h in headers if h.lower() == "date"), "Date")
                date_val = str(denorm.get(date_key, "")).strip()
                if not date_val:
                    sl_val = str(denorm.get(sl_key, "")).strip() if sl_key else ""
                    desc_val = str(denorm.get("Description", denorm.get("Particulars", ""))).strip()
                    lines = p_text.splitlines()
                    # 1. Match exact Sl number at start of line
                    if sl_val and sl_val.isdigit():
                        for l in lines:
                            if re.match(r"^\s*" + re.escape(sl_val) + r"\b", l):
                                m_d = re.search(r"\b(\d{2}[/-]\d{2}[/-]\d{4})\b", l)
                                if m_d:
                                    date_val = m_d.group(1)
                                    break
                    # 2. Match price and description words in line
                    if not date_val:
                        for l in lines:
                            if (str(int(denorm.get("Gross Amount", denorm.get("NetAmt", -1)))) in l or str(int(denorm.get("Rate", denorm.get("Price", -1)))) in l) and any(w.lower() in l.lower() for w in desc_val.split() if len(w) > 3):
                                m_d = re.search(r"\b(\d{2}[/-]\d{2}[/-]\d{4})\b", l)
                                if m_d:
                                    date_val = m_d.group(1)
                                    break
                    # 3. Fallback to last seen date
                    if not date_val and last_seen_date:
                        date_val = last_seen_date
                if date_val:
                    last_seen_date = date_val
                    denorm[date_key] = date_val

            denorm_items.append(denorm)
        p["items"] = denorm_items

        denorm_taxes = []
        for tax in p.get("taxes", []):
            denorm_taxes.append(denormalize_item(tax, headers))
        p["taxes"] = denorm_taxes

    return {
        "document_name": filename,
        "total_pages": total_pages,
        "total_items": total_items,
        "pages": pages_results
    }

# ─────────────────────────────────────────────────────────────────────────────
#  FILE I/O
# ─────────────────────────────────────────────────────────────────────────────

def save_json(data: dict, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as fh:
        json.dump(data, fh, indent=2, ensure_ascii=False)
    log.info("  [OUT] Saved → %s", output_path)


import csv as _csv

def save_csv(data: dict, output_path: Path) -> None:
    pages = data.get("pages", [])
    
    # Gather all unique keys across all items on all pages

    all_keys = set()
    all_items = []
    for page in pages:
        items = page.get("items", []) + page.get("taxes", [])
        for item in items:
            # Inject page_number so it's in the CSV
            item["_page_number"] = page.get("page_number")
            all_keys.update(item.keys())
            all_items.append(item)
            
    if not all_keys:
        log.info("  [CSV] No items to save.")
        return
        
    # Sort fieldnames logically: _page_number first, then others
    fieldnames = ["_page_number"] + sorted([k for k in all_keys if k != "_page_number"])
    
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", newline="", encoding="utf-8") as fh:
        writer = _csv.DictWriter(fh, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(all_items)
    log.info("  [CSV] Saved → %s", output_path)


def save_excel(data: dict, output_path: Path) -> None:
    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Invoice Items"
        
        pages = data.get("pages", [])
        all_keys = set()
        all_items = []
        for page in pages:
            for item in (page.get("items", []) + page.get("taxes", [])):
                item["_page_number"] = page.get("page_number")
                all_keys.update(item.keys())
                all_items.append(item)
        if not all_items:
            return
            
        preferred_order = ["_page_number", "Particulars", "Quantity", "Price", "NetAmt", "Gross Amount", "Rate"]
        fieldnames = [k for k in preferred_order if k in all_keys] + [k for k in sorted(all_keys) if k not in preferred_order]
        
        ws.append(fieldnames)
        for row in all_items:
            ws.append([row.get(k, "") for k in fieldnames])
            
        wb.save(output_path)
        log.info("  [XLSX] Saved → %s", output_path)
    except ImportError:
        pass


def append_to_master_excel(data: dict, master_path: Path) -> None:
    try:
        import openpyxl
        master_path.parent.mkdir(parents=True, exist_ok=True)
        if master_path.exists():
            wb = openpyxl.load_workbook(master_path)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Master Billing Ledger"
            ws.append(["Document Name", "Page Number", "Particulars", "Quantity", "Price", "NetAmt"])
            
        doc_name = data.get("document_name", "")
        for page in data.get("pages", []):
            page_num = page.get("page_number", "")
            for item in (page.get("items", []) + page.get("taxes", [])):
                part = item.get("Particulars", item.get("description", item.get("Name", "")))
                qty = item.get("Quantity", item.get("Qty", item.get("Count", 1.0)))
                price = item.get("Price", item.get("Rate", item.get("Amount", 0.0)))
                net = item.get("NetAmt", item.get("Total", price))
                ws.append([doc_name, page_num, part, qty, price, net])
                
        wb.save(master_path)
        log.info("  [MASTER LEDGER] Appended invoice rows → %s", master_path)
    except ImportError:
        pass


def stitch_cross_page_splits(pages_text: list[str]) -> list[str]:
    header_keywords = [
        r"manipal hospital",
        r"survey no",
        r"cin:",
        r"inpatient interim bill",
        r"date wise itemised bill",
        r"name\s*:\s*\w+",
        r"age/sex\s*:",
        r"inpatient no\s*:",
        r"reg no\s*:",
        r"sl\.\s*particulars",
        r"order item\s+qty",
        r"location\s*:",
    ]
    footer_keywords = [
        r"page\s+\d+\s+of\s+\d+",
        r"\b\d+\s+of\s+\d+\b",
        r"gross total\s*:",
        r"grand total\s*:",
        r"sub total\s*:",
    ]
    
    def is_header_or_footer(line: str) -> bool:
        line_lower = line.lower().strip()
        if not line_lower:
            return True
        if re.match(r'^[\s\-_=_]*$', line_lower):
            return True
        for pat in header_keywords + footer_keywords:
            if re.search(pat, line_lower):
                return True
        return False

    def ends_with_price_not_date(line: str) -> bool:
        line_clean = line.strip()
        if not line_clean:
            return False
        # Exclude dates / timestamps
        if re.search(r'\b\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\b', line_clean):
            return False
        if re.search(r'\b\d{2}:\d{2}\s*(?:AM|PM)?\b', line_clean):
            return False
        return bool(re.search(r'\b\d+(?:\.\d{2})?\s*$', line_clean))
        
    stitched_pages = list(pages_text)
    
    for i in range(len(stitched_pages) - 1):
        curr_text = stitched_pages[i]
        next_text = stitched_pages[i+1]
        
        curr_lines = curr_text.splitlines()
        next_lines = next_text.splitlines()
        
        curr_data_idx = -1
        for idx in range(len(curr_lines) - 1, -1, -1):
            if not is_header_or_footer(curr_lines[idx]):
                curr_data_idx = idx
                break
                
        if curr_data_idx == -1:
            continue
            
        curr_last_line = curr_lines[curr_data_idx].strip()
        has_price_curr = ends_with_price_not_date(curr_last_line)
        
        # Check if the split happened
        if not has_price_curr:
            # We want to identify if the next page starts with pricing/date details for the split item
            lines_to_move = []
            next_data_idx = -1
            
            # Find the first data line on the next page
            for idx in range(len(next_lines)):
                if not is_header_or_footer(next_lines[idx]):
                    next_data_idx = idx
                    break
                    
            if next_data_idx == -1:
                continue
                
            # Collect all lines from the top of the next page that belong to the split item.
            # This includes lines that are dates, prices, locations, or short numbers,
            # but stops when we hit a line that has a standard description (letters > 4).
            idx = next_data_idx
            while idx < len(next_lines):
                if is_header_or_footer(next_lines[idx]):
                    idx += 1
                    continue
                    
                line_to_check = next_lines[idx].strip()
                
                # Check if it has a price or is a date or is just numbers/locations
                has_price = ends_with_price_not_date(line_to_check)
                is_date = bool(re.search(r'\b\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\b', line_to_check))
                
                clean_text = re.sub(r'\b(?:charged|packed|returned|location|pharmacy|ward|bed|mvd|mvb|c\s+wing|d\s+wing)\b', '', line_to_check, flags=re.IGNORECASE)
                clean_text = re.sub(r'\d{1,2}[/-]\d{1,2}[/-]\d{2,4}', '', clean_text)
                clean_text = re.sub(r'[^a-zA-Z]', '', clean_text)
                is_price_only = len(clean_text) < 4
                
                if has_price or is_date or is_price_only:
                    lines_to_move.append((idx, next_lines[idx]))
                    idx += 1
                else:
                    break
            
            if lines_to_move:
                # Log what we are moving
                for orig_idx, line_val in lines_to_move:
                    log.info(f"  [STITCHING] Page split detected! Moving line '{line_val.strip()}' from top of Page {i+2} to bottom of Page {i+1}")
                    curr_lines.insert(curr_data_idx + 1, line_val)
                    curr_data_idx += 1
                    
                # Remove moved lines from the next page in reverse order to keep indices correct
                for orig_idx, _ in reversed(lines_to_move):
                    next_lines.pop(orig_idx)
                    
                stitched_pages[i] = "\n".join(curr_lines)
                stitched_pages[i+1] = "\n".join(next_lines)
    return stitched_pages


# ─────────────────────────────────────────────────────────────────────────────
#  MAIN PIPELINE
# ─────────────────────────────────────────────────────────────────────────────


def process_file(file_path: Path) -> None:
    import time
    start_time = time.perf_counter()
    log.info("Processing: %s", file_path.name)

    pages_text = extract_text(file_path)
    if not pages_text:
        elapsed = time.perf_counter() - start_time
        log.warning("  Skipping %s — no text could be extracted (took %.2fs).", file_path.name, elapsed)
        return

    pages_text = stitch_cross_page_splits(pages_text)

    result = extract_with_llm(pages_text, file_path.name, file_path=file_path)

    if result is None:
        elapsed = time.perf_counter() - start_time
        log.warning("  Skipping %s — LLM extraction returned nothing (took %.2fs).", file_path.name, elapsed)
        return

    elapsed_seconds = time.perf_counter() - start_time
    if isinstance(result, dict):
        result["processing_time_seconds"] = round(elapsed_seconds, 2)
    log.info("  [TIMING] Completed %s in %.2f seconds (%.2f minutes).", file_path.name, elapsed_seconds, elapsed_seconds / 60.0)

    output_path = OUTPUT_DIR / (file_path.stem + ".json")
    save_json(result, output_path)

    csv_path = OUTPUT_DIR / (file_path.stem + ".csv")
    save_csv(result, csv_path)

    # xlsx_path = OUTPUT_DIR / (file_path.stem + ".xlsx")
    # save_excel(result, xlsx_path)

    master_xlsx = OUTPUT_DIR / "Master_Hospital_Billing.xlsx"
    append_to_master_excel(result, master_xlsx)


def run_on_paths(file_paths: list[Path]) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    valid: list[Path] = []
    for p in file_paths:
        if not p.exists():
            log.error("File not found: %s", p)
        elif p.is_dir():
            for sub_f in sorted(p.glob("*.*")):
                if sub_f.is_file() and sub_f.suffix.lower() in SUPPORTED_EXTENSIONS:
                    valid.append(sub_f)
        elif not p.is_file():
            log.error("Not a file or directory: %s", p)
        elif p.suffix.lower() not in SUPPORTED_EXTENSIONS:
            log.error(
                "Unsupported file type '%s' for %s", p.suffix, p.name
            )
        else:
            valid.append(p)

    if not valid:
        log.error("No valid files to process. Exiting.")
        return

    log.info("Processing %d file(s).", len(valid))
    success, failed = 0, 0

    for file_path in valid:
        try:
            process_file(file_path)
            success += 1
        except Exception as exc:
            log.error("Unhandled error processing %s: %s", file_path.name, exc)
            failed += 1

    log.info("Done. %d succeeded, %d failed.", success, failed)

# ─────────────────────────────────────────────────────────────────────────────
#  ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Invoice/receipt extractor using Ollama LLM.")
    parser.add_argument("files", nargs="*", type=Path, help="Document files to process")
    parser.add_argument("--model", default=OLLAMA_MODEL, help=f"Ollama model (default: {OLLAMA_MODEL})")
    parser.add_argument("--output-dir", default=str(OUTPUT_DIR), type=Path)
    parser.add_argument("--engine", default="ollama", choices=["ollama", "gemini", "openai"], help="AI Engine selection")
    parser.add_argument("--fallback-engine", default="ollama", choices=["ollama", "gemini", "openai"], help="Fallback AI engine when primary cloud engine fails")
    parser.add_argument("--workers", type=int, default=MAX_WORKERS, help=f"Number of concurrent page threads (default: {MAX_WORKERS})")
    parser.add_argument("--watch", type=Path, help="Folder watchdog mode — continuously monitor folder for new documents")
    args = parser.parse_args()

    OLLAMA_MODEL    = args.model
    OUTPUT_DIR      = args.output_dir
    AI_ENGINE       = args.engine
    FALLBACK_ENGINE = args.fallback_engine
    MAX_WORKERS     = args.workers

    if args.watch:
        import time
        watch_dir = args.watch
        watch_dir.mkdir(parents=True, exist_ok=True)
        log.info("  [WATCHDOG] Continuous Monitoring ACTIVE on folder: '%s'", watch_dir)
        processed = set(p for p in watch_dir.glob("*.*"))
        try:
            while True:
                for f in watch_dir.glob("*.*"):
                    if f.is_file() and f.suffix.lower() in SUPPORTED_EXTENSIONS and f not in processed:
                        log.info("  [WATCHDOG] Incoming document detected: %s", f.name)
                        process_file(f)
                        processed.add(f)
                time.sleep(2)
        except KeyboardInterrupt:
            log.info("  [WATCHDOG] Stopped monitoring.")
    elif args.files:
        run_on_paths(args.files)
    else:
        default_dir = Path("docs") if Path("docs").exists() else Path(".")
        found_files = sorted([f for f in default_dir.glob("*.*") if f.is_file() and f.suffix.lower() in SUPPORTED_EXTENSIONS])
        if found_files:
            log.info("No files specified. Automatically scanning '%s' folder...", default_dir)
            run_on_paths(found_files)
        else:
            parser.print_help()
